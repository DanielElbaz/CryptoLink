"""
CryptoLink - Exchange file import and cataloging tool
Run with: python main.py
Then open: http://127.0.0.1:5000
"""

from flask import Flask, request, jsonify, send_file
import pandas as pd
import uuid
import io
import os
import re
import sys
import json
import time
import atexit
import threading
import webbrowser
import traceback
from datetime import datetime

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # keep UTF-8 (Hebrew, etc.) readable in JSON responses


def _bundled_resource_path(relative_path):
    """Path to a read-only resource bundled with the app (e.g. static/vis-network.min.js).
    When frozen into a standalone .exe by PyInstaller, such files are unpacked into a
    temporary directory at runtime (sys._MEIPASS) rather than living next to the script."""
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def _persistent_data_dir():
    """Directory for files that must survive between runs (the autosave). When frozen,
    sys._MEIPASS (used above for bundled resources) is wiped as soon as the exe exits, so
    anything written there would vanish - this instead resolves next to the .exe itself."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# The Graph tab's vis-network library is embedded directly into the page (see index()
# below) instead of loaded from a CDN or a separate /static request - this used to pull
# from unpkg.com at runtime, which silently breaks the whole tab on any offline/restricted
# network (common for forensic workstations). Vendored file, read once at startup.
# Missing file (e.g. main.py was copied without its static/ folder) degrades to a disabled
# Graph tab instead of crashing the whole app on startup.
try:
    with open(_bundled_resource_path(os.path.join("static", "vis-network.min.js")), encoding="utf-8") as _f:
        VIS_NETWORK_JS = _f.read()
except FileNotFoundError:
    print("WARNING: static/vis-network.min.js not found - the Graph tab will be disabled. "
          "Copy the whole project (including the static/ folder) to fix this.")
    VIS_NETWORK_JS = "console.error('vis-network.min.js missing - copy the static/ folder from the project.');"

# In-memory storage (no database yet)
SUSPECTS = {}   # suspect_id -> {"name": str}
CATALOG = {}    # file_id -> {..., "suspect_id": str, "dataframe": DataFrame}
# lower(address) -> original-case address, manually marked as noise (e.g. a known exchange
# hot wallet touched by unrelated users) and excluded from address-based cross-referencing
# (Wallets/Transfers/Graph). Individual transactions through it still show in Amounts.
EXCLUDED_ADDRESSES = {}
# lower(address) -> {"address": original-case address, "note": free text}, e.g. "confirmed
# via subpoena" or "known legit exchange wallet, not suspicious". Purely informational -
# doesn't affect matching, filtering, or the graph.
ADDRESS_NOTES = {}

# Free-text label for the investigation currently open, e.g. "Case 2026-03 - Maharan".
# Purely a display/filename convenience - shown in the header and used as the export
# filename, carried over by Save/Load case, reset to "" by Clean all.
CASE_NAME = ""

# lower(address) -> {"address": original-case address, "sightings": [{suspect_name,
# case_label, exchanges, occurrence_count, first_seen, last_seen, added_at}, ...]}.
# A persistent, growing ledger of every wallet ever committed from a finished
# investigation - unlike everything else above, it survives "Load case" (a case swaps out
# the working set; this is knowledge carried between cases). Populated only by the explicit
# "Add wallets to database" action, never automatically, so it doesn't fill up with
# still-unverified data from a case in progress.
KNOWN_WALLETS = {}

# ---------------------------------------------------------------------------
# DETECTION
# ---------------------------------------------------------------------------

EXCHANGE_KEYWORDS = {
    "binance": ["binance"],
    "okx": ["okx", "okex"],
    "bybit": ["bybit"],
    "redotpay": ["redot", "redotpay", "redot pay"],
    "matrix": ["matrix"],
}

NORMALIZATION_SCHEMAS = {
    ("binance", "deposit"): {
        "date": "Create Time", "currency": "Currency", "amount": "Amount",
        "amount_usd": "USDT",
        "network": "Network", "txid": "TXID",
        "external_address": "Source Address", "exchange_address": "Deposit Address",
        "user_id": "User ID", "status": "Status",
    },
    ("binance", "withdrawal"): {
        "date": "Apply Time", "currency": "Currency", "amount": "Amount",
        "amount_usd": "USDT",
        "network": "Network", "txid": "txId",
        "external_address": "Destination Address", "exchange_address": None,
        "user_id": "User ID", "status": "Status",
    },
    ("okx", "deposit"): {
        "date": "creation time", "currency": "currency", "amount": "amount",
        "network": None, "txid": "txid",
        "external_address": None, "exchange_address": "address",
        "user_id": "uuid", "status": None,
    },
    ("okx", "withdrawal"): {
        "date": "creation time", "currency": "currency", "amount": "amount",
        "network": None, "txid": "txid",
        "external_address": "address", "exchange_address": None,
        "user_id": "uuid", "status": None,
    },
    ("bybit", "deposit"): {
        "date": "deposit_coin_time", "currency": "coin", "amount": "change_amount",
        "amount_usd": "change_amount_usd",
        "network": None, "txid": "tx_id",
        "external_address": "from_address", "exchange_address": "to_address",
        "user_id": "user_id", "status": None,
    },
    ("bybit", "withdrawal"): {
        "date": "submitted_time", "currency": "coin", "amount": "amount",
        "amount_usd": "amount_usd",
        "network": None, "txid": "tx_id",
        "external_address": "to_address", "exchange_address": "from_address",
        "user_id": "user_id", "status": None,
    },
    ("matrix", "deposit"): {
        "date": "Time", "currency": "Currency", "amount": "Amount",
        "network": "Coin", "txid": "Transaction hash",
        "external_address": "Originating Address", "exchange_address": "Recharge address",
        "user_id": None, "status": None,
    },
    ("matrix", "withdrawal"): {
        "date": "Time", "currency": "Currency", "amount": "Amount",
        "network": "Blockchain", "txid": "Transaction hash",
        "external_address": "Withdrawal address", "exchange_address": None,
        "user_id": None, "status": None,
    },
    ("redotpay", "deposit"): {
        "date": "Time", "currency": "Currency", "amount": "Amount",
        "network": "Blockchain", "txid": "Hash",
        "external_address": "Originating Address", "exchange_address": "Deposit Address",
        "user_id": None, "status": None,
    },
    ("redotpay", "withdrawal"): {
        "date": "Time", "currency": "Currency", "amount": "Amount",
        "network": "Blockchain", "txid": "Hash",
        "external_address": "Withdrawal Address", "exchange_address": None,
        "user_id": None, "status": None,
    },
}


def normalize_dataframe(df, exchange, file_type):
    schema = NORMALIZATION_SCHEMAS.get((exchange, file_type))
    if not schema:
        return None, ["no_schema_for_this_exchange"]

    col_lookup = {str(c).strip().lower(): c for c in df.columns}
    normalized = pd.DataFrame(index=df.index)
    missing = []

    for field, source_col in schema.items():
        if source_col is None:
            normalized[field] = None
            continue
        actual_col = col_lookup.get(source_col.strip().lower())
        if actual_col is None:
            normalized[field] = None
            missing.append(field)
        else:
            normalized[field] = df[actual_col]

    return normalized, missing


DEPOSIT_KEYWORDS = ["deposit", "depot", "recu", "receive", "top up", "topup"]
WITHDRAWAL_KEYWORDS = ["withdraw", "retrait", "send", "payout"]


_ALL_KNOWN_COLUMNS = {
    v.strip().lower()
    for schema in NORMALIZATION_SCHEMAS.values()
    for v in schema.values() if v
}


def find_header_row(raw_df, max_scan_rows=15):
    """Some exchange exports have metadata/KYC rows above the real column
    titles (e.g. case info, account holder name) before the actual header row.
    Scans the first rows for the one that best matches known column names
    across all exchanges, and returns its 0-indexed position. Falls back to
    row 0 (the normal case) if nothing better is found."""
    best_row_idx, best_score = 0, 0
    scan_limit = min(max_scan_rows, len(raw_df))
    for i in range(scan_limit):
        row_values = {str(v).strip().lower() for v in raw_df.iloc[i].tolist() if pd.notna(v)}
        score = len(row_values & _ALL_KNOWN_COLUMNS)
        if score > best_score:
            best_score, best_row_idx = score, i
    return best_row_idx if best_score >= 2 else 0


def parse_sheet_smart(excel_file, sheet_name):
    """Parses a sheet, auto-detecting the real header row even if metadata
    rows (KYC info, case references, etc.) precede it."""
    raw = excel_file.parse(sheet_name, header=None)
    header_idx = find_header_row(raw)
    return excel_file.parse(sheet_name, header=header_idx)


def detect_exchange(filename, sheet_name, columns):
    for exchange, keywords in EXCHANGE_KEYWORDS.items():
        for kw in keywords:
            if kw in filename.lower() or kw in sheet_name.lower():
                return exchange

    # Fallback: match the uploaded file's actual column names against the real
    # column names captured in NORMALIZATION_SCHEMAS (from real exchange files),
    # instead of guessed keywords. Far more reliable when filename/sheet name
    # carries no brand hint (e.g. a generic export filename).
    col_set = {str(c).strip().lower() for c in columns}
    best_match, best_score = None, 0
    for (exchange, _file_type), schema in NORMALIZATION_SCHEMAS.items():
        expected_cols = {v.strip().lower() for v in schema.values() if v}
        score = len(col_set & expected_cols)
        if score > best_score:
            best_score, best_match = score, exchange

    # Require at least 2 matching columns - a single shared generic column
    # name (e.g. just "Currency" or "Amount") isn't a reliable signal on its own.
    return best_match if best_score >= 2 else "unknown"


def detect_file_type(filename, sheet_name, columns):
    text_blob = " ".join([filename.lower(), sheet_name.lower()])
    has_deposit = any(kw in text_blob for kw in DEPOSIT_KEYWORDS)
    has_withdrawal = any(kw in text_blob for kw in WITHDRAWAL_KEYWORDS)

    if has_deposit and not has_withdrawal:
        return "deposit"
    if has_withdrawal and not has_deposit:
        return "withdrawal"

    col_blob = " ".join(c.lower() for c in columns)
    if any(kw in col_blob for kw in DEPOSIT_KEYWORDS):
        return "deposit"
    if any(kw in col_blob for kw in WITHDRAWAL_KEYWORDS):
        return "withdrawal"

    return None  # not relevant -> will be ignored


# ---------------------------------------------------------------------------
# ROUTES - SUSPECTS
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return HTML_PAGE.replace("%%VIS_NETWORK_JS%%", VIS_NETWORK_JS, 1)


@app.route("/suspects", methods=["GET"])
def list_suspects():
    return jsonify([{"id": sid, "name": s["name"], "note": s.get("note", "")} for sid, s in SUSPECTS.items()])


@app.route("/suspects", methods=["POST"])
def create_suspect():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    sid = str(uuid.uuid4())
    SUSPECTS[sid] = {"name": name, "note": ""}
    return jsonify({"id": sid, "name": name})


@app.route("/suspects/<suspect_id>", methods=["DELETE"])
def delete_suspect(suspect_id):
    SUSPECTS.pop(suspect_id, None)
    to_remove = [fid for fid, f in CATALOG.items() if f["suspect_id"] == suspect_id]
    for fid in to_remove:
        del CATALOG[fid]
    return jsonify({"success": True})


@app.route("/suspects/<suspect_id>/note", methods=["POST"])
def set_suspect_note(suspect_id):
    if suspect_id not in SUSPECTS:
        return jsonify({"error": "Suspect not found"}), 404
    data = request.get_json()
    SUSPECTS[suspect_id]["note"] = (data.get("note") or "").strip()
    return jsonify({"success": True})


@app.route("/addresses/note", methods=["POST"])
def set_address_note():
    data = request.get_json()
    addr = (data.get("address") or "").strip()
    note = (data.get("note") or "").strip()
    if not addr:
        return jsonify({"error": "Address required"}), 400
    if note:
        ADDRESS_NOTES[addr.lower()] = {"address": addr, "note": note}
    else:
        ADDRESS_NOTES.pop(addr.lower(), None)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# ROUTES - FILES
# ---------------------------------------------------------------------------

@app.route("/files", methods=["GET"])
def list_files():
    """Every catalogued file across all suspects, in the same shape /upload returns per file -
    used to repopulate the Files tab client-side after a case is loaded (no re-upload needed)."""
    return jsonify([{
        "id": file_id, "filename": entry["filename"], "sheet_name": entry["sheet_name"],
        "exchange": entry["exchange"], "file_type": entry["file_type"],
        "columns": entry["columns"], "row_count": entry["row_count"],
        "suspect_id": entry["suspect_id"], "missing_fields": entry["missing_fields"],
    } for file_id, entry in CATALOG.items()])


@app.route("/upload", methods=["POST"])
def upload():
    files = request.files.getlist("files")
    suspect_id = request.form.get("suspect_id")

    if not suspect_id or suspect_id not in SUSPECTS:
        return jsonify({"error": "No active suspect selected"}), 400
    if not files:
        return jsonify({"error": "No file received"}), 400

    results = []
    ignored_count = 0

    for f in files:
        # Filename may arrive in various encodings depending on browser/OS -
        # normalize defensively so non-Latin filenames (Hebrew, etc.) never crash the request.
        try:
            filename = f.filename or "unnamed_file"
        except Exception:
            filename = "unnamed_file"

        try:
            file_bytes = f.read()
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
        except Exception as e:
            results.append({
                "id": str(uuid.uuid4()), "filename": filename,
                "error": f"Unable to read this file: {str(e)}",
            })
            continue

        for sheet_name in excel_file.sheet_names:
            try:
                df = parse_sheet_smart(excel_file, sheet_name)
            except Exception as e:
                results.append({
                    "id": str(uuid.uuid4()), "filename": filename, "sheet_name": sheet_name,
                    "error": f"Unable to read sheet '{sheet_name}': {str(e)}",
                })
                continue

            if df.empty or len(df.columns) == 0:
                continue

            columns = [str(c) for c in df.columns]
            file_type = detect_file_type(filename, sheet_name, columns)

            # Only keep Deposit / Withdrawal sheets - everything else is silently ignored
            if file_type is None:
                ignored_count += 1
                continue

            exchange = detect_exchange(filename, sheet_name, columns)
            file_id = str(uuid.uuid4())
            normalized_df, missing_fields = normalize_dataframe(df, exchange, file_type)

            CATALOG[file_id] = {
                "filename": filename, "sheet_name": sheet_name,
                "exchange": exchange, "file_type": file_type,
                "columns": columns, "row_count": len(df),
                "dataframe": df, "suspect_id": suspect_id,
                "normalized_dataframe": normalized_df, "missing_fields": missing_fields,
            }

            results.append({
                "id": file_id, "filename": filename, "sheet_name": sheet_name,
                "exchange": exchange, "file_type": file_type,
                "columns": columns, "row_count": len(df), "suspect_id": suspect_id,
                "missing_fields": missing_fields,
            })

    return jsonify({"results": results, "ignored_count": ignored_count})


@app.route("/update_label", methods=["POST"])
def update_label():
    data = request.get_json()
    file_id = data.get("id")
    field = data.get("field")
    value = data.get("value")

    if file_id not in CATALOG:
        return jsonify({"error": "File not found"}), 404
    if field not in ("exchange", "file_type", "suspect_id"):
        return jsonify({"error": "Invalid field"}), 400

    CATALOG[file_id][field] = value
    return jsonify({"success": True})


@app.route("/preview/<file_id>")
def preview(file_id):
    if file_id not in CATALOG:
        return jsonify({"error": "File not found"}), 404

    try:
        entry = CATALOG[file_id]
        df = entry["dataframe"].head(15)
        # fillna("") before stringifying: pandas' astype(str) leaves empty cells
        # as raw NaN instead of converting them to text, which breaks JSON output
        # (this is exactly what was happening with MatrixPort files with blank cells)
        preview_data = df.fillna("").astype(str).to_dict(orient="records")
        return jsonify({
            "filename": entry["filename"], "sheet_name": entry["sheet_name"],
            "columns": entry["columns"], "rows": preview_data,
        })
    except Exception as e:
        # Surface the real error to the UI instead of a silent/generic failure,
        # so any issue tied to specific file content can be diagnosed precisely.
        traceback.print_exc()
        return jsonify({"error": f"Preview failed: {str(e)}"}), 500


def _to_float(val):
    try:
        if val is None:
            return None
        f = float(val)
        return None if f != f else f  # filters NaN
    except (TypeError, ValueError):
        return None


def _to_datetime(val):
    try:
        ts = pd.to_datetime(val, errors="coerce")
        return None if pd.isna(ts) else ts
    except Exception:
        return None


def _clean_str(val):
    """Converts a pandas cell to a clean string, or None if empty/NaN.
    Needed because pandas leaves blank cells as raw float NaN, which
    breaks JSON output (same root cause as the earlier MatrixPort preview bug -
    here applied to every text field, not just external_address)."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s == "" or s.lower() == "nan" else s


def get_all_transactions(suspect_ids=None):
    """Flattens every imported sheet's normalized data into one list of dicts,
    each tagged with which suspect/exchange/file it came from.
    If suspect_ids is given (a set/list), only those suspects' data is included -
    used to scope exports to selected suspects."""
    rows = []
    for file_id, entry in CATALOG.items():
        if suspect_ids is not None and entry["suspect_id"] not in suspect_ids:
            continue
        ndf = entry.get("normalized_dataframe")
        if ndf is None:
            continue
        suspect_name = SUSPECTS.get(entry["suspect_id"], {}).get("name", "Unknown suspect")
        has_usd = "amount_usd" in ndf.columns

        for _, r in ndf.iterrows():
            rows.append({
                "file_id": file_id,
                "suspect_id": entry["suspect_id"],
                "suspect_name": suspect_name,
                "exchange": entry["exchange"],
                "file_type": entry["file_type"],
                "date": _to_datetime(r.get("date")),
                "currency": _clean_str(r.get("currency")),
                "amount": _to_float(r.get("amount")),
                "amount_usd": _to_float(r.get("amount_usd")) if has_usd else None,
                "txid": _clean_str(r.get("txid")),
                "external_address": _clean_str(r.get("external_address")),
                # The exchange's own side of the address pair (e.g. Matrix's "Recharge
                # address"/"Withdrawal address" is external_address's counterpart on that same
                # row) - purely informational, never used for suspect/wallet matching.
                "exchange_address": _clean_str(r.get("exchange_address")),
            })
    return rows


def _address_note(address):
    return ADDRESS_NOTES.get((address or "").lower(), {}).get("note", "")


def compute_addresses_analysis(suspect_ids=None):
    """Dominant wallets + addresses shared across different suspect/exchange combos."""
    rows = [r for r in get_all_transactions(suspect_ids)
            if r["external_address"] and r["external_address"].lower() not in EXCLUDED_ADDRESSES]

    # Group case-insensitively (ETH checksummed vs lowercase should still match),
    # but keep the first-seen original casing for display.
    by_address = {}
    for r in rows:
        by_address.setdefault(r["external_address"].lower(), []).append(r)

    results = []
    for addr_lower, occurrences in by_address.items():
        distinct_combos = {(o["suspect_id"], o["exchange"]) for o in occurrences}
        distinct_suspects = {o["suspect_id"] for o in occurrences}
        results.append({
            "address": occurrences[0]["external_address"],
            "note": _address_note(addr_lower),
            # Sightings of this address from previously committed cases (see
            # /known_wallets/commit) - a wallet with almost no data in the current case can
            # still surface a match against everything ever investigated before.
            "known_sightings": KNOWN_WALLETS.get(addr_lower, {}).get("sightings", []),
            # Free-text label (e.g. "Binance hot wallet", "Mixer") + optional badge color, set
            # manually per-wallet and stored in the same persistent cross-case ledger as
            # sightings - so it's shown next to this address everywhere it recurs, in this
            # case and any future one.
            "category": KNOWN_WALLETS.get(addr_lower, {}).get("category", ""),
            "category_color": KNOWN_WALLETS.get(addr_lower, {}).get("category_color", ""),
            "occurrence_count": len(occurrences),
            "distinct_accounts": len(distinct_combos),
            "is_cross_account": len(distinct_combos) > 1,
            # Distinguishes "same person, multiple exchanges" (cross_account but not
            # cross_suspect) from "different people" (cross_suspect) - the two are very
            # different findings and were previously conflated into one flag.
            "distinct_suspect_count": len(distinct_suspects),
            "is_cross_suspect": len(distinct_suspects) > 1,
            "occurrences": [{
                "suspect_id": o["suspect_id"], "suspect_name": o["suspect_name"], "exchange": o["exchange"],
                "file_type": o["file_type"], "amount": o["amount"], "amount_usd": o["amount_usd"],
                "currency": o["currency"],
                "date": o["date"].isoformat() if o["date"] is not None else None,
                "txid": o["txid"], "exchange_address": o["exchange_address"],
                "exchange_address_category": KNOWN_WALLETS.get((o["exchange_address"] or "").lower(), {}).get("category", ""),
                "exchange_address_category_color": KNOWN_WALLETS.get((o["exchange_address"] or "").lower(), {}).get("category_color", ""),
                "exchange_address_note": _address_note(o["exchange_address"]),
            } for o in occurrences],
        })

    results.sort(key=lambda x: x["occurrence_count"], reverse=True)
    return results


def _parse_suspect_ids():
    """Reads the optional ?suspects=id1,id2 query param shared by every analysis/export
    route, so the same suspect-filter selection drives both the on-screen tabs and exports."""
    suspects_param = request.args.get("suspects", "").strip()
    return set(suspects_param.split(",")) if suspects_param else None


@app.route("/analysis/addresses")
def analysis_addresses():
    return jsonify(compute_addresses_analysis(_parse_suspect_ids()))


@app.route("/addresses/excluded", methods=["GET"])
def list_excluded_addresses():
    """Hidden wallets with enough context to decide whether to restore one without having
    to go re-check the Wallets tab - just the raw address isn't enough to remember why it
    mattered."""
    rows = [r for r in get_all_transactions() if r["external_address"]]
    by_address = {}
    for r in rows:
        by_address.setdefault(r["external_address"].lower(), []).append(r)

    result = []
    for addr_lower, original in EXCLUDED_ADDRESSES.items():
        occurrences = by_address.get(addr_lower, [])
        result.append({
            "address": original,
            "occurrence_count": len(occurrences),
            "suspect_names": sorted({o["suspect_name"] for o in occurrences}),
            "note": _address_note(addr_lower),
        })
    result.sort(key=lambda x: x["address"].lower())
    return jsonify(result)


@app.route("/addresses/exclude", methods=["POST"])
def exclude_address():
    data = request.get_json()
    addr = (data.get("address") or "").strip()
    if not addr:
        return jsonify({"error": "Address required"}), 400
    EXCLUDED_ADDRESSES[addr.lower()] = addr
    return jsonify({"success": True})


@app.route("/addresses/include", methods=["POST"])
def include_address():
    data = request.get_json()
    addr = (data.get("address") or "").strip().lower()
    EXCLUDED_ADDRESSES.pop(addr, None)
    return jsonify({"success": True})


def compute_amounts_analysis(suspect_ids=None):
    """Every transaction, sorted largest to smallest (USD-equivalent first when available)."""
    rows = get_all_transactions(suspect_ids)

    def sort_key(r):
        # rows with a USD value sort among themselves by that value;
        # rows without one fall back to raw amount, sorted after.
        return (r["amount_usd"] is not None, r["amount_usd"] or r["amount"] or 0)

    rows.sort(key=sort_key, reverse=True)

    def _cat(address):
        return KNOWN_WALLETS.get((address or "").lower(), {})

    return [{
        "suspect_name": r["suspect_name"], "exchange": r["exchange"], "file_type": r["file_type"],
        "amount": r["amount"], "amount_usd": r["amount_usd"], "currency": r["currency"],
        "date": r["date"].isoformat() if r["date"] is not None else None,
        "external_address": r["external_address"], "exchange_address": r["exchange_address"], "txid": r["txid"],
        # Same persistent cross-case category lookup used by compute_addresses_analysis - a
        # wallet tagged from any tab shows the same badge here too. Both addresses on the row
        # can be tagged independently (e.g. the external wallet AND the exchange's own address).
        "category": _cat(r["external_address"]).get("category", ""),
        "category_color": _cat(r["external_address"]).get("category_color", ""),
        "exchange_address_category": _cat(r["exchange_address"]).get("category", ""),
        "exchange_address_category_color": _cat(r["exchange_address"]).get("category_color", ""),
        # Same free-text note lookup used by compute_addresses_analysis - a note set from any
        # tab shows up here too.
        "note": _address_note(r["external_address"]),
        "exchange_address_note": _address_note(r["exchange_address"]),
    } for r in rows]


@app.route("/analysis/amounts")
def analysis_amounts():
    return jsonify(compute_amounts_analysis(_parse_suspect_ids()))


def _transfer_side(o):
    known = KNOWN_WALLETS.get((o["external_address"] or "").lower(), {})
    exchange_known = KNOWN_WALLETS.get((o["exchange_address"] or "").lower(), {})
    return {
        "suspect_id": o["suspect_id"], "suspect_name": o["suspect_name"], "exchange": o["exchange"],
        "amount": o["amount"], "amount_usd": o["amount_usd"], "currency": o["currency"],
        "date": o["date"].isoformat(), "txid": o["txid"],
        "external_address": o["external_address"], "exchange_address": o["exchange_address"],
        "category": known.get("category", ""), "category_color": known.get("category_color", ""),
        "exchange_address_category": exchange_known.get("category", ""),
        "exchange_address_category_color": exchange_known.get("category_color", ""),
        "note": _address_note(o["external_address"]),
        "exchange_address_note": _address_note(o["exchange_address"]),
    }


# Hard requirements for an ADDRESS-based (heuristic) match - unlike a TXID match (proof),
# address reuse alone is circumstantial and needs corroboration to count as a real transfer
# instead of coincidental/unrelated reuse of the same wallet months apart.
ADDRESS_MATCH_AMOUNT_TOLERANCE = 0.05   # max 5% relative difference when both amounts are known
ADDRESS_MATCH_MAX_GAP_HOURS = 7 * 24    # max 7 days between the withdrawal and the deposit


def _amount_rel_diff(w, d):
    """Relative amount difference between a withdrawal and a deposit, or None if they're not
    in the same currency or either side's amount is missing (nothing to compare)."""
    same_currency = (
        w["currency"] is not None and d["currency"] is not None
        and w["currency"].strip().upper() == d["currency"].strip().upper()
    )
    if same_currency and w["amount"] and d["amount"]:
        return abs(w["amount"] - d["amount"]) / max(abs(w["amount"]), abs(d["amount"]), 1e-9)
    return None


def _address_match_ok(w, d):
    """Gate for ADDRESS-based candidates only (see compute_transfers_analysis pass 2): reject
    pairs more than 7 days apart, and reject pairs whose amounts are both known but differ by
    more than 5% - either one on its own means the address getting reused is most likely
    unrelated reuse (e.g. a suspect's own wallet used for many separate transactions), not
    this specific transfer."""
    gap_hours = (d["date"] - w["date"]).total_seconds() / 3600
    if gap_hours > ADDRESS_MATCH_MAX_GAP_HOURS:
        return False
    rel_diff = _amount_rel_diff(w, d)
    if rel_diff is not None and rel_diff > ADDRESS_MATCH_AMOUNT_TOLERANCE:
        return False
    return True


def _pair_score(w, d):
    """Ranks a candidate (withdrawal, deposit) match sharing the same external address.
    Lower is better. Same-currency pairs whose amounts are close (tier 0) always outrank
    pairs with mismatched/unknown currency or amount (tier 1, only compared by time gap) -
    this is what lets the 1-to-1 matching below pick the single most plausible deposit for
    each withdrawal instead of pairing every withdrawal with every later deposit."""
    gap_hours = (d["date"] - w["date"]).total_seconds() / 3600
    rel_diff = _amount_rel_diff(w, d)
    tier = 0 if rel_diff is not None else 1
    return (tier, rel_diff if rel_diff is not None else 0.0, gap_hours)


def compute_transfers_analysis(suspect_ids=None):
    """Links a withdrawal to a deposit two ways, in priority order:

    1. TXID match (confirmed): the exact same blockchain transaction hash was logged by
       both sides - one exchange recorded it as an outgoing withdrawal, another (or the same
       one) recorded it as an incoming deposit. Same hash = definitely the same transfer, no
       guessing. This also catches cases where one side's export has no usable address column
       at all but does log the tx hash.
    2. Address match (probable): no shared txid, but the withdrawal's destination address
       equals a later deposit's source address. Heuristic - only accepted within 7 days and
       (when the amount is known on both sides) within 5% of each other (see
       _address_match_ok); ranked among surviving candidates by amount closeness then time
       gap. Without these gates, any two unrelated transactions that happened to reuse the
       same wallet months apart and for wildly different amounts would get shown as a
       "transfer" just because nothing better was available for that address.

    Every row is matched to AT MOST ONE counterpart: txid matching runs first and claims
    what it can with certainty, then address matching runs on whatever's left, greedily
    pairing best-match-first. This avoids pairing every withdrawal on an address with every
    later deposit on it (the earlier bug, which turned e.g. 3 withdrawals + 4 deposits on one
    address into up to 12 fabricated "transfers").

    Returns {"matched": [...], "unmatched": [...]}. Unmatched entries are withdrawals/deposits
    that had a txid or address to work with but couldn't be confidently linked to anything -
    still worth a human's attention, listed separately rather than dropped or force-matched.
    """
    rows = [r for r in get_all_transactions(suspect_ids)
            if r["date"] is not None and (r["txid"] or r["external_address"])]
    claimed = [False] * len(rows)
    matched = []

    def emit(wi, di, match_type):
        w, d = rows[wi], rows[di]
        tier, rel_diff, _ = _pair_score(w, d)
        claimed[wi] = claimed[di] = True
        matched.append({
            "address": w["external_address"] or d["external_address"],
            "match_type": match_type,
            "withdrawal": _transfer_side(w),
            "deposit": _transfer_side(d),
            "gap_hours": round(abs((d["date"] - w["date"]).total_seconds()) / 3600, 2),
            "same_suspect": w["suspect_id"] == d["suspect_id"],
            "same_exchange": w["exchange"] == d["exchange"],
            "amount_matched": tier == 0,
            "amount_diff_pct": round(rel_diff * 100, 2) if tier == 0 else None,
        })

    # Pass 1: exact TXID match - highest confidence, runs first so it claims rows before
    # the address heuristic gets a chance to guess wrong.
    by_txid = {}
    for i, r in enumerate(rows):
        if r["txid"]:
            by_txid.setdefault(r["txid"].strip().lower(), []).append(i)
    for idxs in by_txid.values():
        withdrawals = [i for i in idxs if rows[i]["file_type"] == "withdrawal"]
        deposits = [i for i in idxs if rows[i]["file_type"] == "deposit"]
        candidates = sorted(
            (abs((rows[di]["date"] - rows[wi]["date"]).total_seconds()), wi, di)
            for wi in withdrawals for di in deposits
        )
        for _, wi, di in candidates:
            if not claimed[wi] and not claimed[di]:
                emit(wi, di, "txid")

    # Pass 2: shared external address (only rows the txid pass didn't already claim; manually
    # excluded/noise addresses never participate in address-based matching).
    by_address = {}
    for i, r in enumerate(rows):
        if not claimed[i] and r["external_address"] and r["external_address"].lower() not in EXCLUDED_ADDRESSES:
            by_address.setdefault(r["external_address"].lower(), []).append(i)
    for idxs in by_address.values():
        withdrawals = [i for i in idxs if rows[i]["file_type"] == "withdrawal"]
        deposits = [i for i in idxs if rows[i]["file_type"] == "deposit"]
        candidates = []
        for wi in withdrawals:
            for di in deposits:
                if rows[di]["date"] < rows[wi]["date"]:
                    continue  # deposit happened before the withdrawal - not this direction
                if not _address_match_ok(rows[wi], rows[di]):
                    continue  # too far apart in time or amount - likely unrelated reuse of the address
                candidates.append((_pair_score(rows[wi], rows[di]), wi, di))
        candidates.sort(key=lambda c: c[0])
        for score, wi, di in candidates:
            if not claimed[wi] and not claimed[di]:
                emit(wi, di, "address")

    unmatched = [
        {"address": r["external_address"], "direction": r["file_type"], **_transfer_side(r)}
        for i, r in enumerate(rows) if not claimed[i]
    ]

    matched.sort(key=lambda p: (p["match_type"] != "txid", p["gap_hours"]))
    unmatched.sort(key=lambda u: u["date"], reverse=True)
    return {"matched": matched, "unmatched": unmatched}


@app.route("/analysis/transfers")
def analysis_transfers():
    return jsonify(compute_transfers_analysis(_parse_suspect_ids()))


# Same reasoning as ADDRESS_MATCH_MAX_GAP_HOURS: chaining two already-detected transfers is
# itself a heuristic (the same person moving on to their next transaction), so it gets its
# own gate rather than being assumed indefinitely.
CHAIN_MAX_HOP_GAP_HOURS = 7 * 24         # max 7 days between one hop's deposit and the next hop's withdrawal
CHAIN_AMOUNT_OVERSHOOT_TOLERANCE = 0.05  # the outgoing hop can't exceed the incoming one by more than 5% (fee/rounding slack)
CHAIN_MAX_DEPTH = 12                     # hops per chain - guards against runaway recursion on dense data
CHAIN_MAX_RESULTS = 500                  # total chains returned - same purpose


def _can_chain_hop(p1, p2):
    """Whether transfer p2 plausibly continues transfer p1 - p1's recipient goes on to be
    p2's sender, soon enough after and (when a USD value is known on both sides) without
    sending out more than they'd just received. Not proof of anything, same spirit as the
    address-match heuristic: a trail worth a human's attention, not a certainty."""
    d, w = p1["deposit"], p2["withdrawal"]
    if d["suspect_id"] != w["suspect_id"]:
        return False
    gap_hours = (datetime.fromisoformat(w["date"]) - datetime.fromisoformat(d["date"])).total_seconds() / 3600
    if gap_hours < 0 or gap_hours > CHAIN_MAX_HOP_GAP_HOURS:
        return False
    d_usd, w_usd = d.get("amount_usd"), w.get("amount_usd")
    if d_usd and w_usd and w_usd > d_usd * (1 + CHAIN_AMOUNT_OVERSHOOT_TOLERANCE):
        return False
    return True


def compute_transfer_chains(suspect_ids=None):
    """Links already-detected 1-hop transfers end-to-end into full A -> B -> C -> ... paths:
    money someone received in one transfer, then went on to send in another, shortly after.
    On its own, compute_transfers_analysis only shows isolated 2-party hops - this traces how
    far a trail of money actually goes instead of leaving the reader to connect the dots
    between separate cards by hand.

    Chains fork and join rather than being forced into single linear paths: a transfer can
    have more than one plausible continuation (money getting split between two recipients) or
    predecessor (two deposits both funding one later withdrawal), so the same hop can appear
    in multiple returned chains. Only maximal chains are returned (at least 2 hops, and not a
    sub-path of a longer chain already in the results).
    """
    transfers = compute_transfers_analysis(suspect_ids)["matched"]
    n = len(transfers)

    next_map = [[] for _ in range(n)]
    prev_has_any = [False] * n
    for i in range(n):
        for j in range(n):
            if i != j and _can_chain_hop(transfers[i], transfers[j]):
                next_map[i].append(j)
                prev_has_any[j] = True

    chains = []
    starts = [i for i in range(n) if next_map[i] and not prev_has_any[i]]

    def walk(path, visited):
        if len(chains) >= CHAIN_MAX_RESULTS:
            return
        i = path[-1]
        can_extend = len(path) < CHAIN_MAX_DEPTH
        extended = False
        if can_extend:
            for j in next_map[i]:
                if j in visited:
                    continue  # guards against a cycle, shouldn't happen given the date ordering above
                extended = True
                walk(path + [j], visited | {j})
        if not extended and len(path) > 1:
            chains.append([transfers[k] for k in path])

    for s in starts:
        walk([s], {s})

    chains.sort(key=lambda c: len(c), reverse=True)
    return chains


@app.route("/analysis/chains")
def analysis_chains():
    return jsonify(compute_transfer_chains(_parse_suspect_ids()))


def _fmt_amount_export(amount, currency):
    if amount is None:
        return "-"
    return f"{amount:,.8f} {currency or ''}".rstrip("0").rstrip(".") if amount != int(amount) else f"{int(amount):,} {currency or ''}"


def _fmt_usd_export(amount_usd):
    return f"${amount_usd:,.2f}" if amount_usd is not None else ""


def _fmt_date_export(iso_str):
    if not iso_str:
        return "-"
    try:
        return datetime.fromisoformat(iso_str).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso_str


def build_report_context(suspect_ids=None):
    """Gathers everything an export needs: the 3 analysis datasets plus
    metadata (generation time, suspects included). If suspect_ids is given,
    the report is scoped to only those suspects."""
    scoped_suspects = ([s for sid, s in SUSPECTS.items() if sid in suspect_ids]
                       if suspect_ids is not None else list(SUSPECTS.values()))
    names = sorted({s["name"] for s in scoped_suspects})
    suspect_notes = sorted(
        ({"name": s["name"], "note": s["note"]} for s in scoped_suspects if s.get("note")),
        key=lambda x: x["name"].lower()
    )
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "suspect_names": names or ["(none)"],
        "suspect_notes": suspect_notes,
        "addresses": compute_addresses_analysis(suspect_ids),
        "amounts": compute_amounts_analysis(suspect_ids),
        "transfers": compute_transfers_analysis(suspect_ids),
        "chains": compute_transfer_chains(suspect_ids),
    }


# ---------------------------------------------------------------------------
# EXPORT - Excel
# ---------------------------------------------------------------------------

def export_xlsx(suspect_ids=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    ctx = build_report_context(suspect_ids)
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def write_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    # --- Suspects sheet ---
    ws0 = wb.active
    ws0.title = "Suspects"
    write_header(ws0, ["Name", "Note"])
    for name in ctx["suspect_names"]:
        note = next((n["note"] for n in ctx["suspect_notes"] if n["name"] == name), "")
        ws0.append([name, note])

    # --- Wallets sheet (one row per occurrence, address repeated) ---
    ws = wb.create_sheet("Wallets")
    write_header(ws, ["Address", "Note", "Seen In Other Cases", "Total Occurrences", "Distinct Accounts", "Different People",
                       "Suspect", "Exchange", "Type", "Amount", "Currency", "Amount USD", "Date", "TXID"])
    for item in ctx["addresses"]:
        known_summary = "; ".join(f"{s['case_label']} ({s['suspect_name']})" for s in item["known_sightings"])
        for o in item["occurrences"]:
            ws.append([
                item["address"], item["note"], known_summary, item["occurrence_count"], item["distinct_accounts"],
                "YES" if item["is_cross_suspect"] else "",
                o["suspect_name"], o["exchange"], o["file_type"],
                o["amount"], o["currency"], o["amount_usd"], _fmt_date_export(o["date"]), o["txid"],
            ])

    # --- Amounts sheet ---
    ws2 = wb.create_sheet("Amounts")
    write_header(ws2, ["Suspect", "Exchange", "Type", "Amount", "Currency", "Amount USD",
                        "Date", "External Address", "TXID"])
    for r in ctx["amounts"]:
        ws2.append([
            r["suspect_name"], r["exchange"], r["file_type"], r["amount"], r["currency"],
            r["amount_usd"], _fmt_date_export(r["date"]), r["external_address"], r["txid"],
        ])

    # --- Transfers sheet (one confirmed withdrawal <-> deposit match per row) ---
    ws3 = wb.create_sheet("Transfers")
    write_header(ws3, ["Match Type", "Address", "Same Person", "Same Exchange", "Amount Match", "Gap (hours)",
                        "Withdrawal Suspect", "Withdrawal Exchange", "Withdrawal Amount", "Withdrawal Date", "Withdrawal TXID",
                        "Deposit Suspect", "Deposit Exchange", "Deposit Amount", "Deposit Date", "Deposit TXID"])
    for p in ctx["transfers"]["matched"]:
        w, d = p["withdrawal"], p["deposit"]
        ws3.append([
            "TXID" if p["match_type"] == "txid" else "Address",
            p["address"], "YES" if p["same_suspect"] else "NO", "YES" if p["same_exchange"] else "NO",
            "YES" if p["amount_matched"] else "unverified", p["gap_hours"],
            w["suspect_name"], w["exchange"], w["amount"], _fmt_date_export(w["date"]), w["txid"],
            d["suspect_name"], d["exchange"], d["amount"], _fmt_date_export(d["date"]), d["txid"],
        ])

    # --- Unmatched movements sheet (shared address, no confident counterpart found) ---
    ws4 = wb.create_sheet("Unmatched movements")
    write_header(ws4, ["Address", "Direction", "Suspect", "Exchange", "Amount", "Currency", "Date", "TXID"])
    for u in ctx["transfers"]["unmatched"]:
        ws4.append([
            u["address"], u["direction"], u["suspect_name"], u["exchange"],
            u["amount"], u["currency"], _fmt_date_export(u["date"]), u["txid"],
        ])

    # --- Chains sheet (one row per hop, chain number repeated) ---
    ws5 = wb.create_sheet("Chains")
    write_header(ws5, ["Chain #", "Hop #", "Match Type", "From Suspect", "From Exchange", "Amount",
                        "Date", "To Suspect", "To Exchange", "Amount", "Date", "Gap (hours)"])
    for chain_num, chain in enumerate(ctx["chains"], start=1):
        for hop_num, p in enumerate(chain, start=1):
            w, d = p["withdrawal"], p["deposit"]
            ws5.append([
                chain_num, hop_num, "TXID" if p["match_type"] == "txid" else "Address",
                w["suspect_name"], w["exchange"], w["amount"], _fmt_date_export(w["date"]),
                d["suspect_name"], d["exchange"], d["amount"], _fmt_date_export(d["date"]), p["gap_hours"],
            ])

    for ws_ in wb.worksheets:
        for col in ws_.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/export/xlsx")
def export_report():
    suspect_ids = _parse_suspect_ids()
    try:
        buf = export_xlsx(suspect_ids)
        return send_file(buf, as_attachment=True, download_name="cryptolink_report.xlsx",
                          mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


def compute_graph_data(suspect_ids=None):
    """Builds a node/edge graph: suspects and wallet addresses as nodes,
    transactions as edges between them (aggregated per suspect-address pair).
    Only addresses that are actually cross-account (is_cross_account) are
    included - i.e. the address links either two different suspects, or the
    same suspect across two different exchanges. A wallet only ever seen on
    one suspect's one exchange isn't a "connection" and adds noise, so it's
    excluded entirely (no toggle - this is the graph's whole purpose)."""
    addresses = compute_addresses_analysis(suspect_ids)
    nodes = {}
    edges = []

    for item in addresses:
        if not item["is_cross_account"]:
            continue

        addr_id = f"addr:{item['address']}"
        nodes[addr_id] = {
            "id": addr_id,
            "label": item["address"][:6] + "…" + item["address"][-4:] if len(item["address"]) > 12 else item["address"],
            "title": item["address"],
            "note": item["note"],
            # Red for a wallet shared between different people, amber when it's the same
            # person reusing a wallet across their own exchange accounts - visually distinct
            # findings, not the same alert level.
            "group": "address_shared_cross_suspect" if item["is_cross_suspect"] else "address_shared_same_suspect",
        }

        by_suspect = {}
        for o in item["occurrences"]:
            key = o["suspect_id"]
            agg = by_suspect.setdefault(key, {
                "suspect_name": o["suspect_name"], "count": 0,
                "deposits": 0, "withdrawals": 0, "total_usd": 0.0, "exchanges": set(),
                "transactions": [],
            })
            agg["count"] += 1
            agg["exchanges"].add(o["exchange"])
            if o["file_type"] == "deposit":
                agg["deposits"] += 1
            else:
                agg["withdrawals"] += 1
            if o["amount_usd"]:
                agg["total_usd"] += o["amount_usd"]
            # Kept so clicking this edge in the Graph tab can list the actual transactions
            # behind it, instead of just the aggregated count/title.
            agg["transactions"].append({
                # o["date"] is already an ISO string here - compute_addresses_analysis
                # converts it before compute_graph_data ever sees these occurrences.
                "exchange": o["exchange"], "file_type": o["file_type"],
                "amount": o["amount"], "amount_usd": o["amount_usd"], "currency": o["currency"],
                "date": o["date"], "txid": o["txid"],
            })

        for suspect_id, agg in by_suspect.items():
            suspect_node_id = f"suspect:{suspect_id}"
            if suspect_node_id not in nodes:
                nodes[suspect_node_id] = {
                    "id": suspect_node_id, "label": agg["suspect_name"],
                    "title": agg["suspect_name"], "group": "suspect",
                }
            title = f"{agg['count']} transaction(s) ({agg['deposits']} deposit, {agg['withdrawals']} withdrawal) via {', '.join(sorted(agg['exchanges']))}"
            if agg["total_usd"]:
                title += f" — ~${agg['total_usd']:,.2f}"
            # Edge is defined suspect -> address. A withdrawal sends money from the suspect to
            # that address (arrow into the address, the "to" end); a deposit receives money
            # from that address into the suspect (arrow into the suspect, the "from" end).
            # Both happening on the same wallet just means both arrowheads are shown.
            arrows = {}
            if agg["withdrawals"] > 0:
                arrows["to"] = {"enabled": True}
            if agg["deposits"] > 0:
                arrows["from"] = {"enabled": True}
            edges.append({
                "from": suspect_node_id, "to": addr_id,
                "value": agg["count"], "title": title, "arrows": arrows,
                "suspect_name": agg["suspect_name"], "address": item["address"],
                "transactions": agg["transactions"],
            })

    # Direct suspect-to-suspect edges for TXID-confirmed transfers between two different
    # people. Needed on top of the address-sharing loop above because a TXID match can link
    # two people even when no usable shared address exists (e.g. one side's export only logs
    # an internal/exchange address, not the counterparty's - see compute_transfers_analysis).
    for p in compute_transfers_analysis(suspect_ids)["matched"]:
        if p["match_type"] != "txid" or p["same_suspect"]:
            continue
        w, d = p["withdrawal"], p["deposit"]
        for suspect_id, name in ((w["suspect_id"], w["suspect_name"]), (d["suspect_id"], d["suspect_name"])):
            node_id = f"suspect:{suspect_id}"
            if node_id not in nodes:
                nodes[node_id] = {"id": node_id, "label": name, "title": name, "group": "suspect"}
        edges.append({
            "from": f"suspect:{w['suspect_id']}", "to": f"suspect:{d['suspect_id']}",
            "value": 3, "color": {"color": "#3ecf8e"}, "arrows": {"to": {"enabled": True}},
            "title": f"Confirmed transfer (same TXID {w['txid']}) — {w['amount']} {w['currency']} on {w['exchange']} → {d['exchange']}",
            "transfer": p,
        })

    return {"nodes": list(nodes.values()), "edges": edges}


@app.route("/analysis/graph")
def analysis_graph():
    return jsonify(compute_graph_data(_parse_suspect_ids()))


@app.route("/delete/<file_id>", methods=["DELETE"])
def delete_file(file_id):
    CATALOG.pop(file_id, None)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# CASE SAVE / LOAD
# ---------------------------------------------------------------------------
# Everything above lives only in memory (SUSPECTS/CATALOG/EXCLUDED_ADDRESSES) - a server
# restart or crash wipes an entire investigation with no way back. These two routes let a
# case be saved to a single JSON file and reloaded later (same machine or a different one),
# preserving suspects, every imported file's data, and hidden-wallet choices.

CASE_FORMAT_VERSION = 1


def _build_case_dict():
    """The full session state as a JSON-safe dict - shared by the manual "Save case"
    download and the periodic background autosave."""
    files = []
    for file_id, entry in CATALOG.items():
        ndf = entry.get("normalized_dataframe")
        files.append({
            "file_id": file_id,
            "filename": entry["filename"], "sheet_name": entry["sheet_name"],
            "exchange": entry["exchange"], "file_type": entry["file_type"],
            "columns": entry["columns"], "row_count": entry["row_count"],
            "suspect_id": entry["suspect_id"], "missing_fields": entry["missing_fields"],
            # pandas' own JSON round-trip handles dates/NaN/dtypes correctly, which a naive
            # to_dict() + json.dumps() would mangle (Timestamps, NaN floats aren't JSON-safe).
            "dataframe_json": entry["dataframe"].to_json(orient="records", date_format="iso"),
            "normalized_dataframe_json": ndf.to_json(orient="records", date_format="iso") if ndf is not None else None,
        })

    return {
        "cryptolink_case_version": CASE_FORMAT_VERSION,
        "saved_at": datetime.now().isoformat(),
        "case_name": CASE_NAME,
        "suspects": SUSPECTS,
        "excluded_addresses": EXCLUDED_ADDRESSES,
        "address_notes": ADDRESS_NOTES,
        "files": files,
    }


def _apply_case_dict(case):
    """Replaces the entire in-memory session with a case dict previously produced by
    _build_case_dict() - shared by manual "Load case" and restoring an autosave. Raises on
    malformed input; the caller decides how to report that."""
    new_catalog = {}
    for entry in case["files"]:
        df = pd.read_json(io.StringIO(entry["dataframe_json"]), orient="records")
        ndf = (pd.read_json(io.StringIO(entry["normalized_dataframe_json"]), orient="records")
               if entry.get("normalized_dataframe_json") else None)
        new_catalog[entry["file_id"]] = {
            "filename": entry["filename"], "sheet_name": entry["sheet_name"],
            "exchange": entry["exchange"], "file_type": entry["file_type"],
            "columns": entry["columns"], "row_count": entry["row_count"],
            "dataframe": df, "suspect_id": entry["suspect_id"],
            "normalized_dataframe": ndf, "missing_fields": entry["missing_fields"],
        }

    # Loading a case replaces the current session entirely - a case file is a full snapshot,
    # not something to merge with whatever's already open.
    global CASE_NAME
    SUSPECTS.clear()
    SUSPECTS.update(case["suspects"])
    CATALOG.clear()
    CATALOG.update(new_catalog)
    EXCLUDED_ADDRESSES.clear()
    EXCLUDED_ADDRESSES.update(case.get("excluded_addresses", {}))
    ADDRESS_NOTES.clear()
    ADDRESS_NOTES.update(case.get("address_notes", {}))
    CASE_NAME = case.get("case_name", "") or ""


@app.route("/case/export")
def export_case():
    case = _build_case_dict()
    buf = io.BytesIO(json.dumps(case, ensure_ascii=False, indent=2).encode("utf-8"))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", CASE_NAME).strip("_")
    prefix = f"cryptolink_{safe_name}" if safe_name else "cryptolink_case"
    return send_file(buf, as_attachment=True, download_name=f"{prefix}_{stamp}.json",
                      mimetype="application/json")


@app.route("/case/name", methods=["GET"])
def get_case_name():
    return jsonify({"name": CASE_NAME})


@app.route("/case/name", methods=["POST"])
def set_case_name():
    global CASE_NAME
    data = request.get_json() or {}
    CASE_NAME = (data.get("name") or "").strip()
    return jsonify({"success": True, "name": CASE_NAME})


@app.route("/case/reset", methods=["POST"])
def reset_case():
    """Clean all: wipes the current investigation (suspects, files, hidden wallets, notes,
    case name) so a new one can start fresh. The persistent cross-case known-wallets ledger
    is untouched - that one is explicitly meant to survive this."""
    global CASE_NAME
    SUSPECTS.clear()
    CATALOG.clear()
    EXCLUDED_ADDRESSES.clear()
    ADDRESS_NOTES.clear()
    CASE_NAME = ""
    if os.path.exists(AUTOSAVE_PATH):
        try:
            os.remove(AUTOSAVE_PATH)
        except Exception:
            traceback.print_exc()
    return jsonify({"success": True})


@app.route("/case/import", methods=["POST"])
def import_case():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file received"}), 400

    try:
        case = json.loads(f.read().decode("utf-8"))
    except Exception as e:
        return jsonify({"error": f"Not a valid CryptoLink case file: {str(e)}"}), 400

    if "suspects" not in case or "files" not in case:
        return jsonify({"error": "Not a valid CryptoLink case file"}), 400

    try:
        _apply_case_dict(case)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to load case: {str(e)}"}), 500

    return jsonify({
        "success": True,
        "suspect_count": len(SUSPECTS),
        "file_count": len(CATALOG),
        "saved_at": case.get("saved_at"),
        "case_name": case.get("case_name", ""),
    })


# ---------------------------------------------------------------------------
# CASE AUTOSAVE
# ---------------------------------------------------------------------------
# A background thread periodically writes the current session straight to a local file (no
# browser download dialog, nothing for the user to click) so a crash or an accidental window
# close loses at most a couple of minutes of work instead of everything since the last manual
# "Save case". Purely a local safety net - it never uploads or sends this anywhere.

AUTOSAVE_PATH = os.path.join(_persistent_data_dir(), "cryptolink_autosave.json")
AUTOSAVE_INTERVAL_SECONDS = 120


def _write_autosave():
    if not SUSPECTS and not CATALOG:
        return  # nothing to protect against losing yet
    try:
        tmp_path = AUTOSAVE_PATH + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(_build_case_dict(), f, ensure_ascii=False)
        os.replace(tmp_path, AUTOSAVE_PATH)  # atomic on both POSIX and Windows - never leaves a half-written file
    except Exception:
        traceback.print_exc()


def _autosave_loop():
    while True:
        time.sleep(AUTOSAVE_INTERVAL_SECONDS)
        _write_autosave()


def start_autosave_thread():
    threading.Thread(target=_autosave_loop, daemon=True).start()


@app.route("/case/autosave/status")
def autosave_status():
    if not os.path.exists(AUTOSAVE_PATH):
        return jsonify({"exists": False})
    try:
        with open(AUTOSAVE_PATH, encoding="utf-8") as f:
            saved_at = json.load(f).get("saved_at")
    except Exception:
        return jsonify({"exists": False})
    return jsonify({"exists": True, "saved_at": saved_at})


@app.route("/case/autosave/restore", methods=["POST"])
def restore_autosave():
    if not os.path.exists(AUTOSAVE_PATH):
        return jsonify({"error": "No autosave found"}), 404
    try:
        with open(AUTOSAVE_PATH, encoding="utf-8") as f:
            case = json.load(f)
        _apply_case_dict(case)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to restore autosave: {str(e)}"}), 500
    return jsonify({
        "success": True,
        "suspect_count": len(SUSPECTS),
        "file_count": len(CATALOG),
        "saved_at": case.get("saved_at"),
        "case_name": case.get("case_name", ""),
    })


# ---------------------------------------------------------------------------
# KNOWN WALLETS (cross-case ledger)
# ---------------------------------------------------------------------------
# Grows across every investigation, not just the one currently open - "Load case" replaces
# the working set (suspects/files/notes) but never touches this. Loaded once at startup and
# saved immediately after each commit (commits are rare, deliberate actions, not worth a
# background thread like the autosave).

KNOWN_WALLETS_PATH = os.path.join(_persistent_data_dir(), "cryptolink_known_wallets.json")


def _load_known_wallets():
    if not os.path.exists(KNOWN_WALLETS_PATH):
        return
    try:
        with open(KNOWN_WALLETS_PATH, encoding="utf-8") as f:
            KNOWN_WALLETS.update(json.load(f))
    except Exception:
        traceback.print_exc()


def _save_known_wallets():
    tmp_path = KNOWN_WALLETS_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(KNOWN_WALLETS, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, KNOWN_WALLETS_PATH)


@app.route("/known_wallets/commit", methods=["POST"])
def commit_known_wallets():
    """Snapshots every wallet in the CURRENT session (all suspects, hidden wallets already
    excluded by compute_addresses_analysis) into the permanent cross-case ledger, tagged with
    a case label so a future match can say where/who it was seen with."""
    data = request.get_json() or {}
    case_label = (data.get("case_label") or "").strip() or datetime.now().strftime("Case %Y-%m-%d %H:%M")
    added_at = datetime.now().isoformat()

    addresses = compute_addresses_analysis(None)
    sighting_count = 0
    for item in addresses:
        addr_lower = item["address"].lower()
        entry = KNOWN_WALLETS.setdefault(addr_lower, {"address": item["address"], "sightings": []})

        by_suspect = {}
        for o in item["occurrences"]:
            agg = by_suspect.setdefault(o["suspect_name"], {"exchanges": set(), "count": 0, "dates": []})
            agg["exchanges"].add(o["exchange"])
            agg["count"] += 1
            if o["date"]:
                agg["dates"].append(o["date"])

        for suspect_name, agg in by_suspect.items():
            entry["sightings"].append({
                "suspect_name": suspect_name, "case_label": case_label,
                "exchanges": sorted(agg["exchanges"]), "occurrence_count": agg["count"],
                "first_seen": min(agg["dates"]) if agg["dates"] else None,
                "last_seen": max(agg["dates"]) if agg["dates"] else None,
                "added_at": added_at,
            })
            sighting_count += 1

    _save_known_wallets()
    return jsonify({
        "success": True, "case_label": case_label,
        "wallet_count": len(addresses), "sighting_count": sighting_count,
    })


@app.route("/known_wallets/category", methods=["POST"])
def set_known_wallet_category():
    """Manually tags a single wallet with a free-text category + optional badge color, stored
    in the same persistent cross-case ledger as commit_known_wallets(). Unlike that bulk
    action, this works one address at a time and doesn't require the address to already have
    any data in the current case - typing in an address here is enough to permanently add it
    to the database (setdefault creates the entry), which is also how a wallet gets added to
    the ledger "manually" without going through a whole case."""
    data = request.get_json() or {}
    address = (data.get("address") or "").strip()
    addr_lower = address.lower()
    if not addr_lower:
        return jsonify({"error": "Address required"}), 400

    category = (data.get("category") or "").strip()
    color = (data.get("color") or "").strip()

    if category or color:
        entry = KNOWN_WALLETS.setdefault(addr_lower, {"address": address, "sightings": []})
        entry["category"] = category
        entry["category_color"] = color
    else:
        # Clearing the category on a wallet that was only ever known for that category (never
        # committed with real sightings) leaves nothing worth keeping - drop it entirely.
        entry = KNOWN_WALLETS.get(addr_lower)
        if entry:
            entry["category"] = ""
            entry["category_color"] = ""
            if not entry.get("sightings"):
                del KNOWN_WALLETS[addr_lower]

    _save_known_wallets()
    return jsonify({"success": True, "address": address, "category": category, "category_color": color})


_load_known_wallets()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

HTML_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CryptoLink - Exchange File Import</title>
<script>%%VIS_NETWORK_JS%%</script>
<style>
    :root {
        --bg: #0f1420; --bg-card: #171d2e; --border: #2a3348;
        --text: #e6e9f0; --text-dim: #8b93a7; --btn-text: #e6e9f0;
        --accent: #4f8cff; --accent-dim: #2a4a8a;
        --success: #3ecf8e; --warning: #f5a623; --danger: #f0556b;
        --input-bg: #101625; --hover-bg: #1c2438; --dropzone-hover-bg: #1a2236; --nested-bg: #10141f;
        --radius: 10px;
    }
    /* Same variables, light values - toggled via the 🌙/☀️ button (see toggleTheme()), which
       stamps data-theme on <html> and remembers the choice in localStorage. */
    :root[data-theme="light"] {
        --bg: #f4f5f8; --bg-card: #ffffff; --border: #dde1e9;
        --text: #1a1f2b; --text-dim: #656d7c; --btn-text: #ffffff;
        --accent: #2451b8; --accent-dim: #3568d4;
        --success: #1a9d63; --warning: #b06a05; --danger: #d1364f;
        --input-bg: #ffffff; --hover-bg: #eef1f6; --dropzone-hover-bg: #eaf0ff; --nested-bg: #f7f8fb;
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: -apple-system, "Segoe UI", Roboto, sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; }
    .container { max-width: 1440px; margin: 0 auto; padding: 32px 20px 100px; }
    header { margin-bottom: 24px; }
    header h1 { font-size: 26px; margin: 0 0 6px; font-weight: 600; }
    header p { color: var(--text-dim); margin: 0; font-size: 14px; }
    .header-row { display: flex; justify-content: space-between; align-items: flex-start; gap: 16px; flex-wrap: wrap; }
    .case-actions { display: flex; gap: 8px; flex-shrink: 0; }
    .case-name-row { display: flex; align-items: center; gap: 6px; margin: 2px 0 8px; }
    .case-name-badge {
        display: inline-flex; align-items: center; gap: 6px; font-size: 13px; color: var(--text-dim);
        background: var(--bg-card); border: 1px solid var(--border); border-radius: 999px;
        padding: 3px 12px; cursor: pointer;
    }
    .case-name-badge:hover { color: var(--text); border-color: var(--accent-dim); }
    .case-name-badge.untitled { font-style: italic; }
    .case-name-input {
        font-size: 13px; background: var(--input-bg); color: var(--text); border: 1px solid var(--accent);
        border-radius: 999px; padding: 3px 12px; width: 220px;
    }
    .autosave-banner {
        display: flex; align-items: center; justify-content: space-between; gap: 12px;
        background: rgba(79,140,255,0.1); border: 1px solid var(--accent-dim); border-radius: var(--radius);
        padding: 12px 16px; margin-bottom: 20px; font-size: 13px;
    }
    .autosave-banner .actions { display: flex; gap: 8px; flex-shrink: 0; }

    .panel {
        background: var(--bg-card); border: 1px solid var(--border);
        border-radius: var(--radius); padding: 18px 20px; margin-bottom: 20px;
    }
    .panel-title { font-size: 13px; color: var(--text-dim); margin-bottom: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.03em; }

    .suspect-row { display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }
    select, input[type="text"] {
        background: var(--input-bg); color: var(--text); border: 1px solid var(--border);
        border-radius: 6px; padding: 9px 12px; font-size: 14px;
    }
    select:focus, input[type="text"]:focus { outline: none; border-color: var(--accent); }
    #activeSuspectSelect { min-width: 220px; }
    .search-row { display: flex; align-items: center; gap: 8px; margin-bottom: 16px; }
    #analysisSearch { width: 100%; max-width: 420px; font-family: "SF Mono", Consolas, monospace; }
    .search-match-note { font-size: 12px; color: var(--text-dim); margin-left: 2px; }
    mark.search-hit { background: rgba(245,166,35,0.35); color: inherit; border-radius: 2px; padding: 0 1px; }

    .btn {
        background: var(--accent-dim); color: var(--btn-text); border: none;
        padding: 9px 16px; border-radius: 6px; cursor: pointer; font-size: 13.5px;
    }
    .btn:hover { background: var(--accent); }
    .btn.small { padding: 6px 12px; font-size: 12.5px; }
    .btn.danger { background: rgba(240,85,107,0.15); color: var(--danger); }
    .btn.danger:hover { background: rgba(240,85,107,0.3); }
    .btn:disabled { opacity: 0.4; cursor: not-allowed; }

    .dropzone {
        border: 2px dashed var(--border); border-radius: var(--radius);
        padding: 40px 24px; text-align: center; cursor: pointer;
        transition: all 0.15s ease; background: var(--bg-card);
    }
    .dropzone:hover, .dropzone.dragover { border-color: var(--accent); background: var(--dropzone-hover-bg); }
    .dropzone.disabled { opacity: 0.5; cursor: not-allowed; }
    .dropzone .icon { font-size: 30px; margin-bottom: 8px; }
    .dropzone .main-text { font-size: 14.5px; font-weight: 500; }
    .dropzone .sub-text { font-size: 12.5px; color: var(--text-dim); margin-top: 4px; }
    input[type="file"] { display: none; }

    .toast {
        position: fixed; bottom: 24px; right: 24px; background: var(--success);
        color: #0a1f16; padding: 13px 20px; border-radius: 8px; font-size: 13.5px;
        font-weight: 600; box-shadow: 0 6px 20px rgba(0,0,0,0.3);
        transform: translateY(20px); opacity: 0; transition: all 0.25s ease;
        z-index: 200; max-width: 380px;
    }
    .toast.show { transform: translateY(0); opacity: 1; }
    .toast.warn { background: var(--warning); }

    .suspect-block { margin-top: 22px; }
    .suspect-header {
        display: flex; justify-content: space-between; align-items: center;
        padding: 12px 4px; cursor: pointer;
    }
    .suspect-header h2 { font-size: 16px; margin: 0; display: flex; align-items: center; gap: 10px; }
    .suspect-header .name-text { unicode-bidi: isolate; }
    .suspect-header .count-badge {
        background: var(--accent-dim); color: var(--text); font-size: 11.5px;
        padding: 2px 9px; border-radius: 20px; font-weight: 600;
    }
    .suspect-header .chevron { transition: transform 0.15s ease; color: var(--text-dim); }
    .suspect-header .chevron.collapsed { transform: rotate(-90deg); }

    .suspect-note {
        font-size: 12.5px; color: var(--text-dim); padding: 0 4px 10px; cursor: pointer;
        white-space: pre-wrap; word-break: break-word;
    }
    .suspect-note:hover { color: var(--text); }
    .suspect-note .note-placeholder { opacity: 0.6; }
    .suspect-note textarea {
        width: 100%; max-width: 480px; min-height: 60px; background: var(--input-bg); color: var(--text);
        border: 1px solid var(--accent); border-radius: 6px; padding: 8px 10px; font-size: 12.5px;
        font-family: inherit; resize: vertical; cursor: text;
    }
    .note-actions { display: flex; gap: 8px; margin-top: 6px; }

    .table-wrap { background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius); overflow: hidden; }
    table { width: 100%; border-collapse: collapse; font-size: 13.5px; }
    thead th { text-align: left; padding: 11px 14px; color: var(--text-dim); font-weight: 500; border-bottom: 1px solid var(--border); white-space: nowrap; }
    tbody td { padding: 10px 14px; border-bottom: 1px solid var(--border); vertical-align: middle; }
    tbody tr:last-child td { border-bottom: none; }
    tbody tr:hover { background: var(--hover-bg); }
    tbody tr.highlight { animation: flashHighlight 1.8s ease; }
    @keyframes flashHighlight { 0% { background: rgba(62,207,142,0.25); } 100% { background: transparent; } }

    /* Isolate bidirectional text (e.g. Hebrew filenames) so it never scrambles
       neighboring Latin text (sheet names, dashes, etc.) */
    .bidi-safe { unicode-bidi: isolate; direction: auto; display: inline-block; }

    select.unknown { border-color: var(--warning); }
    .badge { display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11.5px; font-weight: 600; }
    .badge.deposit { background: rgba(62,207,142,0.15); color: var(--success); }
    .badge.withdrawal { background: rgba(240,85,107,0.15); color: var(--danger); }

    .empty-state { text-align: center; padding: 40px; color: var(--text-dim); font-size: 14px; }

    .modal-overlay {
        display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
        background: rgba(0,0,0,0.6); z-index: 100; align-items: center; justify-content: center;
    }
    .modal-overlay.active { display: flex; }
    .modal { background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius); width: 90%; max-width: 900px; max-height: 80vh; display: flex; flex-direction: column; }
    .modal-header { padding: 16px 20px; border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; gap: 12px; }
    .modal-header h3 { margin: 0; font-size: 15px; unicode-bidi: isolate; overflow: hidden; text-overflow: ellipsis; }
    .modal-close { cursor: pointer; color: var(--text-dim); font-size: 20px; background: none; border: none; flex-shrink: 0; }
    .modal-body { padding: 12px 20px; overflow: auto; }
    .modal-body table { font-size: 12px; }
    .modal-body th { position: sticky; top: 0; background: var(--bg-card); }
    .modal-body td, .modal-body th { unicode-bidi: isolate; }
    .modal-error { color: var(--danger); font-size: 13.5px; padding: 10px 0; }

    @media (max-width: 700px) {
        .table-wrap { overflow-x: auto; }
        table { min-width: 650px; }
        .suspect-row { flex-direction: column; align-items: stretch; }
        #activeSuspectSelect { min-width: 0; }
    }

    .main-tabs {
        display: flex; gap: 6px; margin-bottom: 20px; border-bottom: 1px solid var(--border);
        padding-bottom: 0;
    }
    .tab-btn {
        background: none; border: none; color: var(--text-dim); font-size: 14px;
        padding: 10px 16px; cursor: pointer; border-bottom: 2px solid transparent;
        margin-bottom: -1px;
    }
    .tab-btn:hover { color: var(--text); }
    .tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); font-weight: 600; }

    .sub-tabs {
        display: flex; gap: 4px; margin-bottom: 18px;
        background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius);
        padding: 5px;
    }
    .sub-tabs .tab-btn {
        flex: 1; text-align: center; font-size: 13.5px; font-weight: 500; padding: 11px 14px;
        border-bottom: none; border-radius: 7px; margin-bottom: 0; white-space: nowrap;
    }
    .sub-tabs .tab-btn:hover { color: var(--text); background: var(--hover-bg); }
    .sub-tabs .tab-btn.active {
        color: var(--btn-text); background: var(--accent-dim); border-bottom: none; font-weight: 600;
    }

    .analysis-loading { text-align: center; padding: 40px; color: var(--text-dim); font-size: 14px; }
    .analysis-empty { text-align: center; padding: 40px; color: var(--text-dim); font-size: 14px; }

    .cross-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; background: rgba(240,85,107,0.18); color: var(--danger); margin-left: 8px;
    }
    .diff-suspect-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; background: rgba(245,166,35,0.18); color: var(--warning);
    }
    .same-person-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; background: rgba(79,140,255,0.18); color: var(--accent); margin-left: 8px;
    }
    .unverified-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; background: rgba(245,166,35,0.18); color: var(--warning); margin-left: 8px;
    }
    .known-elsewhere-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; background: rgba(139,92,246,0.18); color: #a78bfa; margin-left: 8px; cursor: pointer;
    }
    .known-elsewhere-badge:hover { background: rgba(139,92,246,0.3); }
    .known-sightings-box {
        background: rgba(139,92,246,0.08); border: 1px solid rgba(139,92,246,0.3);
        border-radius: 8px; padding: 10px 12px; margin-bottom: 12px;
    }
    .known-sightings-title { font-size: 12px; font-weight: 600; color: #a78bfa; margin-bottom: 8px; }
    .known-sightings-box table { font-size: 12px; }
    .category-badge {
        display: inline-block; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 700; color: #fff; margin-left: 8px; cursor: pointer;
    }
    .category-badge-placeholder {
        display: inline-block; font-size: 11px; color: var(--text-dim); margin-left: 8px;
        cursor: pointer; border: 1px dashed var(--border); border-radius: 20px; padding: 2px 8px;
    }
    .category-badge-placeholder:hover { color: var(--text); border-color: var(--accent-dim); }
    .category-edit-row { display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }
    .category-edit-row input[type="text"] { flex: 1 1 120px; min-width: 0; }
    .category-edit-row input[type="color"] { width: 36px; height: 30px; padding: 2px; cursor: pointer; flex-shrink: 0; }
    .tag-preset-row { display: flex; align-items: center; gap: 6px; margin-bottom: 8px; flex-wrap: wrap; }
    .tag-preset-label { font-size: 11px; color: var(--text-dim); margin-right: 2px; }
    .tag-preset-swatch {
        width: 20px; height: 20px; border-radius: 50%; border: 2px solid var(--border);
        cursor: pointer; padding: 0; flex-shrink: 0;
    }
    .tag-preset-swatch:hover { border-color: var(--text); transform: scale(1.12); }
    .note-badge {
        display: inline-block; max-width: 220px; padding: 3px 9px; border-radius: 20px; font-size: 11px;
        font-weight: 600; color: var(--text); background: var(--nested-bg); border: 1px solid var(--border);
        margin-left: 8px; cursor: pointer; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
        vertical-align: middle;
    }
    .note-badge:hover { border-color: var(--accent-dim); }
    .note-badge-placeholder {
        display: inline-block; font-size: 11px; color: var(--text-dim); margin-left: 8px;
        cursor: pointer; border: 1px dashed var(--border); border-radius: 20px; padding: 2px 8px;
    }
    .note-badge-placeholder:hover { color: var(--text); border-color: var(--accent-dim); }
    .note-popover {
        position: fixed; z-index: 60; background: var(--bg-card); border: 1px solid var(--border);
        border-radius: 8px; padding: 10px; box-shadow: 0 8px 24px rgba(0,0,0,0.4); width: 300px;
        max-width: calc(100vw - 24px);
    }
    .note-popover textarea {
        width: 100%; min-height: 80px; background: var(--input-bg); color: var(--text);
        border: 1px solid var(--accent); border-radius: 6px; padding: 8px 10px; font-size: 12.5px;
        font-family: inherit; resize: vertical; cursor: text;
    }
    .note-popover-actions { display: flex; gap: 8px; margin-top: 8px; }
    #knownWalletsPanel { max-width: 560px; }
    .category-popover {
        position: fixed; z-index: 60; background: var(--bg-card); border: 1px solid var(--border);
        border-radius: 8px; padding: 10px; box-shadow: 0 8px 24px rgba(0,0,0,0.4); width: 300px;
        max-width: calc(100vw - 24px);
    }
    .category-popover-actions { display: flex; gap: 8px; margin-top: 8px; }
    .flow-line { display: flex; align-items: center; gap: 6px; font-size: 12px; white-space: nowrap; }
    .flow-line + .flow-line { margin-top: 4px; }
    .flow-tag {
        font-size: 9.5px; font-weight: 700; color: var(--text-dim); background: var(--nested-bg);
        border-radius: 4px; padding: 1px 5px; flex-shrink: 0;
    }
    .flow-arrow { color: var(--text-dim); font-size: 12px; text-align: center; margin: 2px 0; }

    .addr-row { cursor: pointer; }
    .addr-detail-row td { background: var(--nested-bg); padding: 0; }
    .addr-detail-wrap { padding: 10px 30px; }
    .addr-detail-wrap table { font-size: 12.5px; }
    .addr-mono { font-family: "SF Mono", Consolas, monospace; font-size: 12.5px; unicode-bidi: isolate; }

    .transfer-card {
        background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius);
        padding: 16px; margin-bottom: 12px;
    }
    .transfer-card .addr-line { font-size: 12px; color: var(--text-dim); margin-bottom: 10px; }
    .transfer-flow { display: flex; align-items: center; gap: 14px; flex-wrap: wrap; }
    .transfer-side { flex: 1; min-width: 200px; }
    .transfer-side .label { font-size: 11px; color: var(--text-dim); text-transform: uppercase; margin-bottom: 4px; }
    .transfer-side .exchange-name { font-weight: 700; text-transform: capitalize; }
    .transfer-side .amount { font-size: 15px; font-weight: 600; margin-top: 2px; }
    .transfer-side .txid-line { font-size: 11.5px; color: var(--text-dim); margin-top: 8px; word-break: break-all; }
    .transfer-arrow { color: var(--accent); font-size: 20px; }
    .transfer-gap { font-size: 11.5px; color: var(--text-dim); margin-top: 10px; }

    .chain-card {
        background: var(--bg-card); border: 1px solid var(--accent-dim); border-radius: var(--radius);
        padding: 16px; margin-bottom: 20px;
    }
    .chain-header { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }
    .chain-path { font-size: 14px; }
    .chain-card .transfer-card { background: var(--nested-bg); margin-bottom: 0; }
    .chain-link {
        text-align: center; font-size: 12px; color: var(--text-dim); padding: 10px 0;
    }

    .results-note { font-size: 12.5px; color: var(--text-dim); margin-bottom: 12px; }
    .load-more-row { text-align: center; margin-top: 16px; }

    .sortable-th { cursor: pointer; user-select: none; white-space: nowrap; }
    .sortable-th:hover { color: var(--text); }
    .person-filter-row { display: flex; align-items: center; gap: 8px; margin-bottom: 14px; flex-wrap: wrap; }
    .person-filter-row .filter-label { font-size: 12.5px; color: var(--text-dim); margin-right: 2px; }
    .filter-chip {
        background: var(--bg-card); color: var(--text-dim); border: 1px solid var(--border);
        padding: 6px 14px; border-radius: 20px; font-size: 12.5px; cursor: pointer;
    }
    .filter-chip:hover { color: var(--text); border-color: var(--accent); }
    .filter-chip.active { background: var(--accent-dim); color: var(--btn-text); border-color: var(--accent-dim); }

    .export-row { display: flex; align-items: center; gap: 8px; margin-bottom: 16px; flex-wrap: wrap; }
    a.btn { text-decoration: none; display: inline-block; }

    .field-warning {
        display: inline-block; margin-left: 8px; font-size: 11px; font-weight: 600;
        color: var(--warning); cursor: help; border-bottom: 1px dotted var(--warning);
    }

    .export-panel { margin-bottom: 16px; }
    .suspect-filter-panel {
        background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius);
        padding: 12px 14px; margin-top: 8px; max-width: 420px;
    }
    .suspect-filter-actions { display: flex; gap: 8px; margin-bottom: 10px; }
    .suspect-filter-list { display: flex; flex-direction: column; gap: 6px; max-height: 200px; overflow-y: auto; }
    .suspect-filter-item { display: flex; align-items: center; gap: 8px; font-size: 13px; }
    .suspect-filter-item input { cursor: pointer; }

    .analysis-filter-panel {
        background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius);
        padding: 12px 14px; margin-bottom: 16px;
    }
    .filter-panel-row { display: flex; flex-direction: column; align-items: stretch; gap: 10px; margin-bottom: 10px; }
    .filter-panel-row:last-of-type { margin-bottom: 0; }
    .filter-panel-row label { display: flex; flex-direction: column; align-items: stretch; gap: 4px; font-size: 12.5px; color: var(--text-dim); }
    .filter-panel-row input[type="text"] {
        width: 100%; font-size: 12.5px; padding: 5px 8px; background: var(--input-bg); color: var(--text);
        border: 1px solid var(--border); border-radius: 6px;
    }
    .date-filter-field { position: relative; }
    .date-filter-field .date-filter-display { cursor: pointer; padding-right: 26px; }
    .date-filter-field .date-filter-icon {
        position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
        font-size: 12px; pointer-events: none; opacity: 0.75;
    }

    .date-picker-popup {
        position: fixed; z-index: 70; width: 240px; background: var(--bg-card); color: var(--text);
        border: 1px solid var(--border); border-radius: var(--radius); padding: 10px;
        box-shadow: 0 10px 28px rgba(0,0,0,0.35);
    }
    .dp-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 8px; }
    .dp-title { font-size: 12.5px; font-weight: 600; }
    .dp-nav {
        background: none; border: 1px solid var(--border); color: var(--text); cursor: pointer;
        width: 24px; height: 24px; border-radius: 6px; line-height: 1; font-size: 13px;
    }
    .dp-nav:hover { border-color: var(--accent); color: var(--accent); }
    .dp-weekdays, .dp-grid { display: grid; grid-template-columns: repeat(7, 1fr); gap: 2px; }
    .dp-weekdays span { text-align: center; font-size: 10.5px; color: var(--text-dim); padding: 2px 0; }
    .dp-day {
        background: none; border: none; color: var(--text); cursor: pointer; font-size: 12px;
        padding: 6px 0; border-radius: 6px; text-align: center;
    }
    .dp-day:hover { background: var(--hover-bg); }
    .dp-day.dp-outside { color: var(--text-dim); opacity: 0.5; }
    .dp-day.dp-today { border: 1px solid var(--accent-dim); }
    .dp-day.dp-selected { background: var(--accent-dim); color: var(--btn-text); font-weight: 600; }
    .dp-footer {
        display: flex; justify-content: space-between; margin-top: 8px; padding-top: 8px;
        border-top: 1px solid var(--border);
    }
    .dp-link {
        background: none; border: none; color: var(--accent); cursor: pointer; font-size: 12px; padding: 2px 4px;
    }
    .dp-link:hover { text-decoration: underline; }
    .filter-exchange-list { display: flex; flex-wrap: wrap; gap: 10px; }
    .filter-exchange-item { display: flex; align-items: center; gap: 5px; font-size: 12.5px; color: var(--text-dim); cursor: pointer; }
    .filter-exchange-item input { cursor: pointer; }
    #analysisFilterToggleBtn.active { background: var(--accent-dim); color: var(--btn-text); }

    .graph-controls {
        display: flex; justify-content: space-between; align-items: center;
        margin-bottom: 12px; flex-wrap: wrap; gap: 10px;
    }
    .graph-toggle { font-size: 12.5px; color: var(--text-dim); display: flex; align-items: center; gap: 6px; cursor: pointer; }
    .graph-legend { display: flex; gap: 14px; flex-wrap: wrap; }
    .legend-item { font-size: 11.5px; color: var(--text-dim); display: flex; align-items: center; gap: 5px; }
    .legend-dot { width: 10px; height: 10px; border-radius: 50%; display: inline-block; }
    #graphCanvas {
        width: 100%; height: 560px; background: var(--bg-card); border: 1px solid var(--border);
        border-radius: var(--radius);
    }
    .graph-canvas-wrap { position: relative; }
    .node-toolbar {
        position: absolute; display: none; gap: 4px; background: var(--bg-card);
        border: 1px solid var(--border); border-radius: 6px; padding: 4px;
        box-shadow: 0 6px 18px rgba(0,0,0,0.35); z-index: 10;
    }
    .graph-edge-detail {
        margin-top: 16px; background: var(--bg-card); border: 1px solid var(--border);
        border-radius: var(--radius); padding: 14px 16px;
    }

    .analysis-layout { display: flex; gap: 20px; align-items: flex-start; }
    .analysis-main { flex: 1; min-width: 0; }
    .analysis-filters-sidebar {
        flex: 0 0 280px; position: sticky; top: 20px; max-height: calc(100vh - 40px); overflow-y: auto;
    }
    .hidden-wallets-sidebar {
        flex: 0 0 260px; background: var(--bg-card); border: 1px solid var(--border);
        border-radius: var(--radius); padding: 14px; position: sticky; top: 20px;
    }
    .hidden-wallets-sidebar h3 { font-size: 12.5px; color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.03em; margin: 0 0 10px; }
    .hidden-wallet-item { border-bottom: 1px solid var(--border); padding: 10px 0; }
    .hidden-wallet-item:last-child { border-bottom: none; padding-bottom: 0; }
    .hidden-wallet-addr { display: flex; align-items: center; gap: 6px; }
    .hidden-wallet-meta { font-size: 11.5px; color: var(--text-dim); margin: 4px 0 8px; }
    @media (max-width: 1150px) {
        .analysis-layout { flex-direction: column; }
        .hidden-wallets-sidebar, .analysis-filters-sidebar { flex: none; width: 100%; position: static; max-height: none; }
    }

    .row-actions { display: inline-flex; gap: 2px; margin-left: 8px; opacity: 0; transition: opacity 0.12s ease; vertical-align: middle; }
    tr:hover .row-actions, .hidden-wallet-item:hover .row-actions, .hidden-wallet-addr:hover .row-actions { opacity: 1; }
    .icon-btn {
        background: none; border: none; cursor: pointer; color: var(--text-dim); font-size: 13px;
        padding: 3px 5px; border-radius: 4px; line-height: 1;
    }
    .icon-btn:hover { color: var(--text); background: var(--hover-bg); }
    .icon-btn.danger:hover { color: var(--danger); background: rgba(240,85,107,0.12); }
</style>
</head>
<body>
<script>
// Applied inline, before the rest of the page renders, so a saved "light" preference doesn't
// flash dark for a frame first.
document.documentElement.setAttribute("data-theme", localStorage.getItem("cryptolink-theme") || "dark");
</script>

<div class="container">
    <header>
        <div class="header-row">
            <div>
                <h1>CryptoLink</h1>
                <div class="case-name-row">
                    <span class="case-name-badge untitled" id="caseNameDisplay" onclick="startEditCaseName()" title="Click to name this case">📁 Untitled case ✏️</span>
                    <input type="text" class="case-name-input" id="caseNameInput" style="display:none;"
                           placeholder="Case name..." maxlength="120"
                           onblur="saveCaseName()" onkeydown="onCaseNameKeydown(event)">
                </div>
                <p>Import Excel files received from exchanges. Only Deposit/Withdrawal sheets are kept.</p>
            </div>
            <div class="case-actions">
                <button class="btn small" id="themeToggleBtn" onclick="toggleTheme()" title="Switch theme">🌙</button>
                <button class="btn small" onclick="saveCase()">💾 Save case</button>
                <button class="btn small" onclick="document.getElementById('caseFileInput').click()">📂 Load case</button>
                <input type="file" id="caseFileInput" accept=".json" style="display:none;" onchange="loadCase(this.files[0])">
                <button class="btn small" id="cleanAllBtn" onclick="cleanAll()" title="Reset the current case (known wallets database is kept)">🧹 Clean all</button>
            </div>
        </div>
    </header>

    <div id="autosaveBanner" class="autosave-banner" style="display:none;"></div>

    <div class="main-tabs">
        <button class="tab-btn active" id="tabBtnFiles" onclick="switchMainTab('files')">Files</button>
        <button class="tab-btn" id="tabBtnAnalysis" onclick="switchMainTab('analysis')">Analysis</button>
    </div>

    <div id="filesView">
        <div class="panel">
            <div class="panel-title">Active suspect</div>
            <div class="suspect-row">
                <select id="activeSuspectSelect" onchange="onSuspectChange()">
                    <option value="">-- No suspect selected --</option>
                </select>
                <input type="text" id="newSuspectName" placeholder="New suspect name">
                <button class="btn" onclick="addSuspect()">+ Add suspect</button>
                <button class="btn danger" id="deleteSuspectBtn" onclick="removeSuspect()" disabled>Delete this suspect</button>
            </div>
        </div>

        <div class="dropzone disabled" id="dropzone">
            <div class="icon">📁</div>
            <div class="main-text" id="dropzoneText">Select an active suspect above first</div>
            <div class="sub-text">.xlsx / .xls - only Deposit/Withdrawal sheets will be imported</div>
            <input type="file" id="fileInput" multiple accept=".xlsx,.xls">
        </div>

        <div id="suspectBlocks"></div>

        <div class="empty-state" id="emptyState">No files imported yet.</div>
    </div>

    <div id="analysisView" style="display:none;">
        <div class="export-panel">
            <div class="export-row">
                <button class="btn small" onclick="triggerExport()">📊 Export to Excel</button>
                <button class="btn small" id="suspectFilterToggleBtn" onclick="toggleSuspectFilterPanel()">Filter suspects ▾</button>
                <button class="btn small" id="knownWalletsToggleBtn" onclick="toggleKnownWalletsPanel()">📚 Add wallets to database</button>
            </div>
            <div id="suspectFilterPanel" class="suspect-filter-panel" style="display:none;">
                <div class="suspect-filter-actions">
                    <button class="btn small" onclick="setAllSuspectFilter(true)">Select all</button>
                    <button class="btn small" onclick="setAllSuspectFilter(false)">Select none</button>
                </div>
                <div id="suspectFilterList" class="suspect-filter-list"></div>
            </div>
            <div id="knownWalletsPanel" class="suspect-filter-panel" style="display:none;">
                <p style="margin:0 0 10px;font-size:12.5px;color:var(--text-dim);">
                    Saves every wallet in the current session to a permanent cross-case database, so a future
                    investigation with little data of its own can still match against it. Give this case a label
                    to remember it by.
                </p>
                <div class="suspect-filter-actions">
                    <input type="text" id="knownWalletsLabel" placeholder="e.g. Q1 2026 fraud case" style="flex:1;">
                    <button class="btn small" onclick="commitKnownWallets()">Add</button>
                </div>
                <p style="margin:14px 0 8px;font-size:12.5px;color:var(--text-dim);border-top:1px solid var(--border);padding-top:12px;">
                    Or tag a single wallet manually - works even for an address with no data in the current case,
                    and adds it to the database right away.
                </p>
                <div class="tag-preset-row">
                    <span class="tag-preset-label">Quick tag:</span>
                    <button type="button" class="tag-preset-swatch" style="background:#8b5cf6;" title="Matrix" onclick="applyTagPreset('manualWalletCategory','manualWalletColor','Matrix','#8b5cf6')"></button>
                    <button type="button" class="tag-preset-swatch" style="background:#d4a017;" title="Binance" onclick="applyTagPreset('manualWalletCategory','manualWalletColor','Binance','#d4a017')"></button>
                    <button type="button" class="tag-preset-swatch" style="background:#22c55e;" title="OKX" onclick="applyTagPreset('manualWalletCategory','manualWalletColor','OKX','#22c55e')"></button>
                    <button type="button" class="tag-preset-swatch" style="background:#dc2626;" title="RedotPay" onclick="applyTagPreset('manualWalletCategory','manualWalletColor','RedotPay','#dc2626')"></button>
                    <button type="button" class="tag-preset-swatch" style="background:#3b82f6;" title="Bybit" onclick="applyTagPreset('manualWalletCategory','manualWalletColor','Bybit','#3b82f6')"></button>
                </div>
                <div class="category-edit-row" style="margin-bottom:8px;">
                    <input type="text" id="manualWalletAddress" placeholder="Wallet address">
                    <input type="text" id="manualWalletCategory" placeholder="Category, e.g. Mixer">
                    <input type="color" id="manualWalletColor" value="#4f8cff" title="Badge color">
                </div>
                <div class="suspect-filter-actions">
                    <button class="btn small" onclick="addManualWalletCategory()">Save wallet</button>
                </div>
            </div>
        </div>
        <div class="analysis-layout">
            <aside class="analysis-filters-sidebar">
                <div class="search-row">
                    <input type="text" id="analysisSearch" placeholder="Search by wallet address or TXID..." oninput="onAnalysisSearch(this.value)">
                    <button class="btn small" id="analysisSearchClear" onclick="clearAnalysisSearch()" style="display:none;">Clear</button>
                </div>
                <button class="btn small" id="analysisFilterToggleBtn" onclick="toggleAnalysisFilterPanel()" style="width:100%;margin-bottom:10px;">🔎 Filters ▾</button>
                <div id="analysisFilterPanel" class="analysis-filter-panel" style="display:none;">
                    <div class="filter-panel-row">
                        <label>From
                            <div class="date-filter-field">
                                <input type="text" id="filterDateFromDisplay" class="date-filter-display" placeholder="jj/mm/aaaa" readonly onclick="openDatePicker('filterDateFrom', this)">
                                <input type="hidden" id="filterDateFrom">
                                <span class="date-filter-icon">📅</span>
                            </div>
                        </label>
                        <label>To
                            <div class="date-filter-field">
                                <input type="text" id="filterDateToDisplay" class="date-filter-display" placeholder="jj/mm/aaaa" readonly onclick="openDatePicker('filterDateTo', this)">
                                <input type="hidden" id="filterDateTo">
                                <span class="date-filter-icon">📅</span>
                            </div>
                        </label>
                        <label>Address contains <input type="text" id="filterAddress" placeholder="0xabc..." oninput="onAnalysisFilterChange()"></label>
                        <label>TXID contains <input type="text" id="filterTxid" placeholder="abc123..." oninput="onAnalysisFilterChange()"></label>
                    </div>
                    <div class="filter-panel-row">
                        <span class="filter-label">Exchange:</span>
                        <div id="filterExchangeList" class="filter-exchange-list"></div>
                    </div>
                    <div class="suspect-filter-actions" style="margin-bottom:0;">
                        <button class="btn small" onclick="clearAnalysisFilters()">Clear filters</button>
                    </div>
                </div>
            </aside>
            <div class="analysis-main">
                <div class="sub-tabs">
                    <button class="tab-btn active" id="subTabAddresses" onclick="switchAnalysisTab('addresses')">Wallets</button>
                    <button class="tab-btn" id="subTabAmounts" onclick="switchAnalysisTab('amounts')">Amounts</button>
                    <button class="tab-btn" id="subTabTransfers" onclick="switchAnalysisTab('transfers')">Transfers</button>
                    <button class="tab-btn" id="subTabChains" onclick="switchAnalysisTab('chains')">Chains</button>
                    <button class="tab-btn" id="subTabGraph" onclick="switchAnalysisTab('graph')">Graph</button>
                </div>
                <div id="analysisContent"></div>
            </div>
            <aside id="hiddenWalletsSidebar" class="hidden-wallets-sidebar" style="display:none;"></aside>
        </div>
    </div>
</div>

<div class="modal-overlay" id="modalOverlay">
    <div class="modal">
        <div class="modal-header">
            <h3 id="modalTitle">Preview</h3>
            <button class="modal-close" onclick="closeModal()">&times;</button>
        </div>
        <div class="modal-body" id="modalBody"></div>
    </div>
</div>

<div class="date-picker-popup" id="datePickerPopup" style="display:none;">
    <div class="dp-header">
        <button type="button" class="dp-nav" onclick="dpChangeMonth(-1)">‹</button>
        <span class="dp-title" id="dpTitle"></span>
        <button type="button" class="dp-nav" onclick="dpChangeMonth(1)">›</button>
    </div>
    <div class="dp-weekdays">
        <span>Mo</span><span>Tu</span><span>We</span><span>Th</span><span>Fr</span><span>Sa</span><span>Su</span>
    </div>
    <div class="dp-grid" id="dpGrid"></div>
    <div class="dp-footer">
        <button type="button" class="dp-link" onclick="dpToday()">Today</button>
        <button type="button" class="dp-link" onclick="dpClear()">Clear</button>
    </div>
</div>

<div class="toast" id="toast"></div>

<div id="categoryPopover" class="category-popover" style="display:none;" onclick="event.stopPropagation()">
    <div class="tag-preset-row">
        <span class="tag-preset-label">Quick tag:</span>
        <button type="button" class="tag-preset-swatch" style="background:#8b5cf6;" title="Matrix" onclick="applyTagPreset('categoryPopoverInput','categoryPopoverColorInput','Matrix','#8b5cf6')"></button>
        <button type="button" class="tag-preset-swatch" style="background:#d4a017;" title="Binance" onclick="applyTagPreset('categoryPopoverInput','categoryPopoverColorInput','Binance','#d4a017')"></button>
        <button type="button" class="tag-preset-swatch" style="background:#22c55e;" title="OKX" onclick="applyTagPreset('categoryPopoverInput','categoryPopoverColorInput','OKX','#22c55e')"></button>
        <button type="button" class="tag-preset-swatch" style="background:#dc2626;" title="RedotPay" onclick="applyTagPreset('categoryPopoverInput','categoryPopoverColorInput','RedotPay','#dc2626')"></button>
        <button type="button" class="tag-preset-swatch" style="background:#3b82f6;" title="Bybit" onclick="applyTagPreset('categoryPopoverInput','categoryPopoverColorInput','Bybit','#3b82f6')"></button>
    </div>
    <div class="category-edit-row">
        <input type="text" id="categoryPopoverInput" placeholder="Category, e.g. Binance hot wallet" maxlength="80">
        <input type="color" id="categoryPopoverColorInput" title="Badge color">
    </div>
    <div class="category-popover-actions">
        <button class="btn small" onclick="saveCategoryPopover()">Save</button>
        <button class="btn small" id="categoryPopoverRemoveBtn" onclick="removeCategoryPopover()" style="display:none;">Remove</button>
        <button class="btn small" onclick="closeCategoryPopover()">Cancel</button>
    </div>
</div>

<div id="notePopover" class="note-popover" style="display:none;" onclick="event.stopPropagation()">
    <textarea id="notePopoverInput" placeholder="Add a note about this wallet..." maxlength="2000"></textarea>
    <div class="note-popover-actions">
        <button class="btn small" onclick="saveNotePopover()">Save</button>
        <button class="btn small" id="notePopoverRemoveBtn" onclick="removeNotePopover()" style="display:none;">Remove</button>
        <button class="btn small" onclick="closeNotePopover()">Cancel</button>
    </div>
</div>

<script>
const EXCHANGES = ["binance", "okx", "bybit", "redotpay", "matrix", "unknown"];
const TYPE_LABELS = {"deposit": "Deposit", "withdrawal": "Withdrawal"};

// Quick-tag swatches (categoryPopover + the manual wallet form) prefill the category text +
// color inputs but don't save on their own - free-text/custom color and Save/Remove stay untouched.
function applyTagPreset(textInputId, colorInputId, name, color) {
    document.getElementById(textInputId).value = name;
    document.getElementById(colorInputId).value = color;
}

let suspects = {};       // id -> name
let suspectNotes = {};   // id -> note text
let catalogData = {};    // file id -> entry
let activeSuspectId = "";
let collapsedSuspects = {};
let editingSuspectNote = null;  // suspect id currently showing its note editor, or null
let currentCaseName = "";

const dropzone = document.getElementById("dropzone");
const fileInput = document.getElementById("fileInput");

function refreshSuspects() {
    fetch("/suspects").then(r => r.json()).then(list => {
        suspects = {};
        list.forEach(s => { suspects[s.id] = s.name; suspectNotes[s.id] = s.note || ""; });
        const sel = document.getElementById("activeSuspectSelect");
        sel.innerHTML = '<option value="">-- No suspect selected --</option>';
        list.forEach(s => {
            const opt = document.createElement("option");
            opt.value = s.id; opt.innerText = s.name;
            if (s.id === activeSuspectId) opt.selected = true;
            sel.appendChild(opt);
        });
        renderAll();
        syncSuspectFilterWithSuspects();
    });
}

function addSuspect() {
    const input = document.getElementById("newSuspectName");
    const name = input.value.trim();
    if (!name) { input.focus(); return; }
    fetch("/suspects", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ name })
    }).then(r => r.json()).then(data => {
        if (data.error) { showToast(data.error, true); return; }
        input.value = "";
        activeSuspectId = data.id;
        refreshSuspects();
        updateDropzoneState();
        showToast("Suspect '" + name + "' created and selected", false);
    });
}

function removeSuspect() {
    if (!activeSuspectId) return;
    const name = suspects[activeSuspectId];
    if (!confirm("Delete '" + name + "' and all imported files?")) return;
    fetch("/suspects/" + activeSuspectId, { method: "DELETE" }).then(() => {
        Object.keys(catalogData).forEach(fid => {
            if (catalogData[fid].suspect_id === activeSuspectId) delete catalogData[fid];
        });
        activeSuspectId = "";
        refreshSuspects();
        updateDropzoneState();
    });
}

function onSuspectChange() {
    activeSuspectId = document.getElementById("activeSuspectSelect").value;
    updateDropzoneState();
}

function updateDropzoneState() {
    const hasActive = !!activeSuspectId;
    dropzone.classList.toggle("disabled", !hasActive);
    document.getElementById("deleteSuspectBtn").disabled = !hasActive;
    document.getElementById("dropzoneText").innerText = hasActive
        ? "Drag " + suspects[activeSuspectId] + "'s files here, or click to browse"
        : "Select an active suspect above first";
}

dropzone.addEventListener("click", () => { if (activeSuspectId) fileInput.click(); });
dropzone.addEventListener("dragover", (e) => { e.preventDefault(); if (activeSuspectId) dropzone.classList.add("dragover"); });
dropzone.addEventListener("dragleave", () => dropzone.classList.remove("dragover"));
dropzone.addEventListener("drop", (e) => {
    e.preventDefault();
    dropzone.classList.remove("dragover");
    if (activeSuspectId) handleFiles(e.dataTransfer.files);
});
fileInput.addEventListener("change", (e) => handleFiles(e.target.files));

function handleFiles(fileList) {
    if (!fileList.length || !activeSuspectId) return;
    const formData = new FormData();
    for (const f of fileList) formData.append("files", f);
    formData.append("suspect_id", activeSuspectId);

    const originalText = document.getElementById("dropzoneText").innerText;
    document.getElementById("dropzoneText").innerText = "Importing...";

    fetch("/upload", { method: "POST", body: formData })
        .then(r => r.json())
        .then(data => {
            document.getElementById("dropzoneText").innerText = originalText;
            if (data.error) { showToast(data.error, true); return; }

            const newIds = [];
            const errors = [];
            data.results.forEach(item => {
                if (item.error) { errors.push((item.filename || "file") + ": " + item.error); return; }
                catalogData[item.id] = item;
                newIds.push(item.id);
            });

            renderAll(newIds);

            if (errors.length) {
                showToast("Error: " + errors.join(" / "), true);
            } else if (newIds.length === 0) {
                showToast("No Deposit/Withdrawal sheet detected in this file(s)" +
                    (data.ignored_count ? " (" + data.ignored_count + " tab(s) ignored)" : ""), true);
            } else {
                let msg = "✓ " + newIds.length + " sheet(s) added to " + suspects[activeSuspectId];
                if (data.ignored_count) msg += " (" + data.ignored_count + " other tab(s) ignored)";
                showToast(msg, false);
            }
        })
        .catch(err => {
            document.getElementById("dropzoneText").innerText = originalText;
            showToast("Network error during import: " + err, true);
        });
    fileInput.value = "";
}

function showToast(msg, isWarn) {
    const t = document.getElementById("toast");
    t.innerText = msg;
    t.classList.toggle("warn", !!isWarn);
    t.classList.add("show");
    clearTimeout(window._toastTimer);
    window._toastTimer = setTimeout(() => t.classList.remove("show"), 4200);
}

// Free-text note attached to a suspect (e.g. "confirmed via subpoena"). Click-to-edit
// inline, no popup - consistent with how hiding a wallet works elsewhere in the app.
function suspectNoteHtml(sid) {
    const note = suspectNotes[sid] || "";
    if (editingSuspectNote === sid) {
        return `<div class="suspect-note" onclick="event.stopPropagation()">
            <textarea id="noteInput-${sid}" placeholder="Add a note about this suspect...">${escapeHtml(note)}</textarea>
            <div class="note-actions">
                <button class="btn small" onclick="saveSuspectNote('${sid}')">Save</button>
                <button class="btn small" onclick="cancelSuspectNoteEdit()">Cancel</button>
            </div>
        </div>`;
    }
    return `<div class="suspect-note" onclick="event.stopPropagation(); startSuspectNoteEdit('${sid}')">
        ${note ? `📝 ${escapeHtml(note)}` : '<span class="note-placeholder">+ Add note</span>'}
    </div>`;
}

function startSuspectNoteEdit(sid) {
    editingSuspectNote = sid;
    renderAll();
    const ta = document.getElementById("noteInput-" + sid);
    if (ta) ta.focus();
}

function cancelSuspectNoteEdit() {
    editingSuspectNote = null;
    renderAll();
}

function saveSuspectNote(sid) {
    const note = document.getElementById("noteInput-" + sid).value.trim();
    fetch(`/suspects/${sid}/note`, {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ note })
    }).then(() => {
        suspectNotes[sid] = note;
        editingSuspectNote = null;
        renderAll();
    });
}

function renderAll(highlightIds) {
    highlightIds = highlightIds || [];
    const container = document.getElementById("suspectBlocks");
    container.innerHTML = "";

    const bySuspect = {};
    Object.keys(catalogData).forEach(id => {
        const item = catalogData[id];
        if (!bySuspect[item.suspect_id]) bySuspect[item.suspect_id] = [];
        bySuspect[item.suspect_id].push(id);
    });

    const suspectIds = Object.keys(bySuspect);
    document.getElementById("emptyState").style.display = suspectIds.length ? "none" : "block";

    suspectIds.forEach(sid => {
        const suspectName = suspects[sid] || "Deleted suspect";
        const ids = bySuspect[sid];
        const exchangeCount = new Set(ids.map(id => catalogData[id].exchange)).size;

        const block = document.createElement("div");
        block.className = "suspect-block";

        const collapsed = !!collapsedSuspects[sid];

        block.innerHTML = `
            <div class="suspect-header" onclick="toggleSuspect('${sid}')">
                <h2><span class="chevron ${collapsed ? 'collapsed' : ''}">▾</span> <span class="name-text">${escapeHtml(suspectName)}</span>
                    <span class="count-badge">${ids.length} sheet(s) - ${exchangeCount} exchange(s)</span>
                </h2>
            </div>
            ${suspectNoteHtml(sid)}
            <div class="table-wrap" style="${collapsed ? 'display:none;' : ''}" id="tw-${sid}">
                <table>
                    <thead>
                        <tr>
                            <th>File</th><th>Sheet</th><th>Exchange</th><th>Type</th><th>Rows</th><th>Actions</th>
                        </tr>
                    </thead>
                    <tbody id="tbody-${sid}"></tbody>
                </table>
            </div>
        `;
        container.appendChild(block);

        const tbody = block.querySelector("#tbody-" + sid);
        ids.forEach(id => {
            const item = catalogData[id];
            const tr = document.createElement("tr");
            if (highlightIds.includes(id)) tr.className = "highlight";
            const missing = item.missing_fields || [];
            const missingWarning = missing.length
                ? `<span class="field-warning" title="Columns not found in this file: ${escapeHtml(missing.join(', '))}. Those fields will be blank for every row - analysis relying on them (e.g. address matching) may miss data here.">⚠ ${missing.length} field(s) not mapped</span>`
                : "";
            tr.innerHTML = `
                <td><span class="bidi-safe">${escapeHtml(item.filename)}</span></td>
                <td><span class="bidi-safe">${escapeHtml(item.sheet_name)}</span></td>
                <td>${buildExchangeSelect(id, item.exchange)}</td>
                <td><span class="badge ${item.file_type}">${TYPE_LABELS[item.file_type]}</span></td>
                <td>${item.row_count}${missingWarning}</td>
                <td>
                    <button class="btn small" onclick="showPreview('${id}')">Preview</button>
                    <button class="btn small danger" onclick="deleteFile('${id}')">Delete</button>
                </td>
            `;
            tbody.appendChild(tr);
        });
    });
}

function toggleSuspect(sid) {
    collapsedSuspects[sid] = !collapsedSuspects[sid];
    renderAll();
}

function buildExchangeSelect(id, current) {
    const cls = current === "unknown" ? "unknown" : "";
    let html = `<select class="${cls}" onclick="event.stopPropagation()" onchange="updateLabel('${id}', 'exchange', this.value, this)">`;
    EXCHANGES.forEach(opt => {
        const label = opt.charAt(0).toUpperCase() + opt.slice(1);
        html += `<option value="${opt}" ${opt === current ? "selected" : ""}>${label}</option>`;
    });
    html += `</select>`;
    return html;
}

function updateLabel(id, field, value, selectEl) {
    fetch("/update_label", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id, field, value })
    }).then(r => r.json()).then(() => {
        catalogData[id][field] = value;
        selectEl.classList.toggle("unknown", value === "unknown");
    });
}

function deleteFile(id) {
    fetch("/delete/" + id, { method: "DELETE" }).then(() => {
        delete catalogData[id];
        renderAll();
    });
}

function showPreview(id) {
    document.getElementById("modalTitle").innerText = "Loading...";
    document.getElementById("modalBody").innerHTML = "";
    document.getElementById("modalOverlay").classList.add("active");

    fetch("/preview/" + id)
        .then(r => r.json().then(data => ({ ok: r.ok, data })))
        .then(({ ok, data }) => {
            if (!ok || data.error) {
                document.getElementById("modalTitle").innerText = "Preview error";
                document.getElementById("modalBody").innerHTML =
                    `<div class="modal-error">${escapeHtml(data.error || "Unknown error")}</div>`;
                return;
            }
            document.getElementById("modalTitle").innerHTML =
                `<span class="bidi-safe">${escapeHtml(data.filename)}</span> - <span class="bidi-safe">${escapeHtml(data.sheet_name)}</span>`;
            let html = "<table><thead><tr>";
            data.columns.forEach(c => html += `<th>${escapeHtml(c)}</th>`);
            html += "</tr></thead><tbody>";
            data.rows.forEach(row => {
                html += "<tr>";
                data.columns.forEach(c => html += `<td>${escapeHtml(String(row[c] ?? ""))}</td>`);
                html += "</tr>";
            });
            html += "</tbody></table>";
            document.getElementById("modalBody").innerHTML = html;
        })
        .catch(err => {
            document.getElementById("modalTitle").innerText = "Preview error";
            document.getElementById("modalBody").innerHTML =
                `<div class="modal-error">Network/parsing error: ${escapeHtml(String(err))}</div>`;
        });
}

function closeModal() { document.getElementById("modalOverlay").classList.remove("active"); }

function escapeHtml(str) {
    const div = document.createElement("div");
    div.innerText = str;
    return div.innerHTML;
}

// ---------------------------------------------------------------------------
// ANALYSIS TAB
// ---------------------------------------------------------------------------

let currentMainTab = "files";
let currentAnalysisTab = "addresses";

function switchMainTab(tab) {
    currentMainTab = tab;
    document.getElementById("filesView").style.display = tab === "files" ? "block" : "none";
    document.getElementById("analysisView").style.display = tab === "analysis" ? "block" : "none";
    document.getElementById("tabBtnFiles").classList.toggle("active", tab === "files");
    document.getElementById("tabBtnAnalysis").classList.toggle("active", tab === "analysis");
    if (tab === "analysis") { loadAnalysisTab(currentAnalysisTab); refreshHiddenWalletsSidebar(); }
}

function switchAnalysisTab(tab) {
    currentAnalysisTab = tab;
    document.getElementById("subTabAddresses").classList.toggle("active", tab === "addresses");
    document.getElementById("subTabAmounts").classList.toggle("active", tab === "amounts");
    document.getElementById("subTabTransfers").classList.toggle("active", tab === "transfers");
    document.getElementById("subTabChains").classList.toggle("active", tab === "chains");
    document.getElementById("subTabGraph").classList.toggle("active", tab === "graph");
    loadAnalysisTab(tab);
}

function loadAnalysisTab(tab) {
    const container = document.getElementById("analysisContent");
    container.innerHTML = '<div class="analysis-loading">Loading...</div>';
    if (tab === "addresses") loadAddresses(container);
    else if (tab === "amounts") loadAmounts(container);
    else if (tab === "transfers") loadTransfers(container);
    else if (tab === "chains") loadChains(container);
    else if (tab === "graph") loadGraph(container);
}

function fmtAmount(amount, currency) {
    if (amount === null || amount === undefined) return "-";
    return amount.toLocaleString(undefined, { maximumFractionDigits: 8 }) + (currency ? " " + currency : "");
}
function fmtUsd(amountUsd) {
    if (amountUsd === null || amountUsd === undefined) return "";
    return "≈ $" + amountUsd.toLocaleString(undefined, { maximumFractionDigits: 2 });
}
function fmtDate(iso) {
    if (!iso) return "-";
    return new Date(iso).toLocaleString(undefined, { hour12: false });
}

// Truncated mono value (address or TXID) with a hover-reveal copy icon that always copies
// the FULL untruncated value - the "…" in the middle was previously copyable as literal
// text along with the two halves, corrupting anything pasted from it.
function truncMono(value, keepStart, keepEnd) {
    keepStart = keepStart || 8;
    keepEnd = keepEnd || 6;
    if (!value) return `<span class="addr-mono">-</span>`;
    const v = String(value);
    const copyBtn = `<span class="row-actions"><button class="icon-btn" title="Copy" onclick="event.stopPropagation(); copyAddress('${escapeHtml(v).replace(/'/g, "\\'")}')">📋</button></span>`;

    // Show the full value (highlighted) instead of truncating when it's what matched the
    // active search - truncation would otherwise hide the very match the user searched for.
    if (analysisSearchQuery && matchesSearch(v)) {
        return `<span class="addr-mono">${highlightMatch(v)}</span>${copyBtn}`;
    }
    if (v.length <= keepStart + keepEnd + 3) {
        return `<span class="addr-mono">${escapeHtml(v)}</span>${copyBtn}`;
    }
    const shortened = v.slice(0, keepStart) + "…" + v.slice(-keepEnd);
    return `<span class="addr-mono" title="${escapeHtml(v)}">${escapeHtml(shortened)}</span>${copyBtn}`;
}

function loadAddresses(container) {
    fetch("/analysis/addresses" + suspectsQueryParam()).then(r => r.json()).then(data => {
        cachedAddresses = data;
        renderAddresses(container);
    }).catch(err => {
        container.innerHTML = `<div class="analysis-empty">Error loading data: ${escapeHtml(String(err))}</div>`;
    });
}

function renderAddresses(container) {
    const data = cachedAddresses || [];
    if (!data.length) {
        container.innerHTML = '<div class="analysis-empty">No addresses found yet - import some files first.</div>';
        return;
    }
    // An address matches if its own text matches, or any of its occurrences' TXID does -
    // the search field covers both wallet addresses and TXIDs per the same box. The
    // structured filter panel additionally requires at least one occurrence to fall within
    // the chosen date range/exchange/address/txid.
    const filtered = data.filter(item =>
        (matchesSearch(item.address) || item.occurrences.some(o => matchesSearch(o.txid))) &&
        item.occurrences.some(o => matchesFilters({ date: o.date, exchange: o.exchange, address: item.address, txid: o.txid }))
    );
    if (!filtered.length) {
        container.innerHTML = `<div class="analysis-empty">No address matches the current search/filters.</div>`;
        return;
    }
    let html = `<div class="results-note">${filtered.length} of ${data.length} distinct address(es) shown, sorted by frequency. Click a row to see all occurrences. Hover an address for actions - wallets that aren't relevant to your case can be hidden from Wallets, Transfers and the Graph.</div>`;
    html += '<div class="table-wrap"><table><thead><tr><th>Address</th><th>Occurrences</th><th>Distinct accounts</th><th></th></tr></thead><tbody>';
    filtered.forEach((item, idx) => {
        const esc = escapeHtml(item.address).replace(/'/g, "\\'");
        const isExpanded = !!analysisSearchQuery || expandedAddresses.has(item.address);
        const sightings = item.known_sightings || [];
        html += `<tr class="addr-row" onclick="toggleAddrDetail('${esc}', ${idx})">
            <td>${addressCellHtml(item.address)}</td>
            <td>${item.occurrence_count}</td>
            <td>${item.distinct_accounts}</td>
            <td>
                ${item.is_cross_suspect ? '<span class="cross-badge">DIFFERENT PEOPLE</span>' : (item.is_cross_account ? '<span class="same-person-badge">SAME PERSON · MULTIPLE EXCHANGES</span>' : "")}
                ${sightings.length ? `<span class="known-elsewhere-badge" onclick="event.stopPropagation(); toggleAddrDetail('${esc}', ${idx})">👁 SEEN ELSEWHERE (${sightings.length})</span>` : ""}
                ${categoryBadgeHtml(item.address, item.category, item.category_color)}
                ${noteBadgeHtml(item.address, item.note)}
            </td>
        </tr>
        <tr class="addr-detail-row" id="addrDetail${idx}" style="display:${isExpanded ? 'table-row' : 'none'};">
            <td colspan="4">
                <div class="addr-detail-wrap">
                    ${sightings.length ? `<div class="known-sightings-box">
                        <div class="known-sightings-title">👁 Seen in ${sightings.length} previous investigation(s)</div>
                        <table><thead><tr><th>Case</th><th>Suspect</th><th>Exchange(s)</th><th>Occurrences</th><th>Last seen</th></tr></thead><tbody>
                        ${sightings.map(s => `<tr>
                            <td>${escapeHtml(s.case_label)}</td>
                            <td>${escapeHtml(s.suspect_name)}</td>
                            <td>${escapeHtml((s.exchanges || []).join(", "))}</td>
                            <td>${s.occurrence_count}</td>
                            <td>${fmtDate(s.last_seen)}</td>
                        </tr>`).join("")}
                        </tbody></table>
                    </div>` : ""}
                    <table><thead><tr><th>Suspect</th><th>Exchange</th><th>Type</th><th>Amount</th><th>Date</th><th>Exchange address</th><th>TXID</th></tr></thead><tbody>
                    ${item.occurrences.map(o => `<tr>
                        <td>${escapeHtml(o.suspect_name)}</td>
                        <td>${escapeHtml(o.exchange)}</td>
                        <td><span class="badge ${o.file_type}">${TYPE_LABELS[o.file_type] || o.file_type}</span></td>
                        <td>${fmtAmount(o.amount, o.currency)}${o.amount_usd ? ' <span style="color:var(--text-dim);font-size:11px;">(' + fmtUsd(o.amount_usd) + ')</span>' : ""}</td>
                        <td>${fmtDate(o.date)}</td>
                        <td>${o.exchange_address ? truncMono(o.exchange_address) + " " + categoryBadgeHtml(o.exchange_address, o.exchange_address_category, o.exchange_address_category_color) + noteBadgeHtml(o.exchange_address, o.exchange_address_note) : '<span style="color:var(--text-dim);">-</span>'}</td>
                        <td class="addr-mono">${highlightMatch(o.txid || "-")}</td>
                    </tr>`).join("")}
                    </tbody></table>
                </div>
            </td>
        </tr>`;
    });
    html += "</tbody></table></div>";
    container.innerHTML = html;
}

// Tracks which addresses are expanded by address (not row index, which shifts whenever the
// list is filtered/re-sorted/re-rendered) so re-rendering after a note edit doesn't collapse
// the row the user was just looking at.
let expandedAddresses = new Set();

function toggleAddrDetail(address, idx) {
    if (expandedAddresses.has(address)) expandedAddresses.delete(address);
    else expandedAddresses.add(address);
    const row = document.getElementById("addrDetail" + idx);
    row.style.display = expandedAddresses.has(address) ? "table-row" : "none";
}

// Address cell with copy/hide icons that only appear on hover (see .row-actions CSS).
// Used anywhere a wallet address is listed. Pass includeHide=false where the wallet is
// already hidden (the sidebar) - a "hide" action there would be redundant with Restore.
function addressCellHtml(address, includeHide) {
    if (includeHide === undefined) includeHide = true;
    const esc = escapeHtml(address).replace(/'/g, "\\'");
    return `<span class="addr-mono">${highlightMatch(address)}</span>
        <span class="row-actions">
            <button class="icon-btn" title="Copy address" onclick="event.stopPropagation(); copyAddress('${esc}')">📋</button>
            ${includeHide ? `<button class="icon-btn danger" title="Hide this wallet" onclick="event.stopPropagation(); hideWallet('${esc}')">🗑️</button>` : ""}
        </span>`;
}

// Category badge/placeholder, usable anywhere a wallet address shows up (Wallets, Amounts,
// Transfers, Chains) - same popover everywhere, so editing one address's category anywhere
// keeps every other view showing it in sync. Values are read back via data-* attributes
// instead of being inlined into the onclick string, so a category or address containing a
// quote can't break the generated HTML.
function categoryBadgeHtml(address, category, color) {
    if (!address) return "";
    const escAddr = escapeHtml(address);
    if (category) {
        return `<span class="category-badge" style="background:${escapeHtml(color || "#4f8cff")};"
            data-address="${escAddr}" data-category="${escapeHtml(category)}" data-color="${escapeHtml(color || "#4f8cff")}"
            onclick="event.stopPropagation(); openCategoryPopoverFromEl(this)" title="Click to edit category">🏷️ ${escapeHtml(category)}</span>`;
    }
    return `<span class="category-badge-placeholder"
        data-address="${escAddr}" data-category="" data-color="#4f8cff"
        onclick="event.stopPropagation(); openCategoryPopoverFromEl(this)">+ Tag</span>`;
}

function openCategoryPopoverFromEl(el) {
    openCategoryPopover(el.dataset.address, el.dataset.category, el.dataset.color, el);
}

function openCategoryPopover(address, category, color, anchorEl) {
    categoryPopoverAddress = address;
    const popover = document.getElementById("categoryPopover");
    document.getElementById("categoryPopoverInput").value = category || "";
    document.getElementById("categoryPopoverColorInput").value = color || "#4f8cff";
    document.getElementById("categoryPopoverRemoveBtn").style.display = category ? "inline-block" : "none";

    const rect = anchorEl.getBoundingClientRect();
    popover.style.display = "block";
    const left = Math.min(rect.left, window.innerWidth - popover.offsetWidth - 12);
    const top = Math.min(rect.bottom + 6, window.innerHeight - popover.offsetHeight - 12);
    popover.style.left = Math.max(12, left) + "px";
    popover.style.top = Math.max(12, top) + "px";

    document.getElementById("categoryPopoverInput").focus();
}

let categoryPopoverAddress = null;

function closeCategoryPopover() {
    document.getElementById("categoryPopover").style.display = "none";
    categoryPopoverAddress = null;
}

// Category/note data lives inside whatever's already cached per tab (cachedAmounts,
// cachedTransfers, etc.) - after an edit, the simplest correct thing is to drop all of it and
// let the current tab re-fetch, rather than trying to patch the same address into every
// possibly-stale array.
function refreshAllCachedAnalysisViews() {
    cachedAddresses = null; cachedAmounts = null; cachedTransfers = null; cachedChains = null;
    loadAnalysisTab(currentAnalysisTab);
}

function saveCategoryPopover() {
    const address = categoryPopoverAddress;
    const category = document.getElementById("categoryPopoverInput").value.trim();
    const color = document.getElementById("categoryPopoverColorInput").value;
    fetch("/known_wallets/category", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address, category, color })
    }).then(() => {
        closeCategoryPopover();
        refreshAllCachedAnalysisViews();
        showToast("Category saved", false);
    }).catch(err => showToast("Network error saving category: " + err, true));
}

function removeCategoryPopover() {
    const address = categoryPopoverAddress;
    fetch("/known_wallets/category", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address, category: "", color: "" })
    }).then(() => {
        closeCategoryPopover();
        refreshAllCachedAnalysisViews();
    }).catch(err => showToast("Network error removing category: " + err, true));
}

document.addEventListener("click", (e) => {
    const popover = document.getElementById("categoryPopover");
    if (popover && popover.style.display !== "none" && !popover.contains(e.target)) closeCategoryPopover();
});

// Free-text note badge/placeholder, usable anywhere a wallet address shows up (Wallets,
// Amounts, Transfers, Chains) - same shared popover everywhere, mirroring categoryBadgeHtml/
// categoryPopover above, so editing a note from any tab keeps every other view in sync.
function noteBadgeHtml(address, note) {
    if (!address) return "";
    const escAddr = escapeHtml(address);
    if (note) {
        const preview = note.length > 40 ? note.slice(0, 40) + "…" : note;
        return `<span class="note-badge"
            data-address="${escAddr}" data-note="${escapeHtml(note)}"
            onclick="event.stopPropagation(); openNotePopoverFromEl(this)" title="${escapeHtml(note)}">📝 ${escapeHtml(preview)}</span>`;
    }
    return `<span class="note-badge-placeholder"
        data-address="${escAddr}" data-note=""
        onclick="event.stopPropagation(); openNotePopoverFromEl(this)">+ Note</span>`;
}

function openNotePopoverFromEl(el) {
    openNotePopover(el.dataset.address, el.dataset.note, el);
}

let notePopoverAddress = null;

function openNotePopover(address, note, anchorEl) {
    notePopoverAddress = address;
    const popover = document.getElementById("notePopover");
    document.getElementById("notePopoverInput").value = note || "";
    document.getElementById("notePopoverRemoveBtn").style.display = note ? "inline-block" : "none";

    const rect = anchorEl.getBoundingClientRect();
    popover.style.display = "block";
    const left = Math.min(rect.left, window.innerWidth - popover.offsetWidth - 12);
    const top = Math.min(rect.bottom + 6, window.innerHeight - popover.offsetHeight - 12);
    popover.style.left = Math.max(12, left) + "px";
    popover.style.top = Math.max(12, top) + "px";

    document.getElementById("notePopoverInput").focus();
}

function closeNotePopover() {
    document.getElementById("notePopover").style.display = "none";
    notePopoverAddress = null;
}

function saveNotePopover() {
    const address = notePopoverAddress;
    const note = document.getElementById("notePopoverInput").value.trim();
    fetch("/addresses/note", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address, note })
    }).then(() => {
        closeNotePopover();
        refreshAllCachedAnalysisViews();
        refreshHiddenWalletsSidebar();
    }).catch(err => showToast("Network error saving note: " + err, true));
}

function removeNotePopover() {
    const address = notePopoverAddress;
    fetch("/addresses/note", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address, note: "" })
    }).then(() => {
        closeNotePopover();
        refreshAllCachedAnalysisViews();
        refreshHiddenWalletsSidebar();
    }).catch(err => showToast("Network error removing note: " + err, true));
}

document.addEventListener("click", (e) => {
    const popover = document.getElementById("notePopover");
    if (popover && popover.style.display !== "none" && !popover.contains(e.target)) closeNotePopover();
});

// Search box shared by Wallets/Amounts/Transfers (Graph handles it separately by
// highlighting/focusing a matching node instead of filtering a list).
let analysisSearchQuery = "";
let cachedAddresses = null, cachedAmounts = null, cachedTransfers = null, cachedChains = null;

// Structured filter panel (Wallets/Amounts/Transfers/Chains) - date range, exchange,
// address substring, TXID substring. Applied client-side on top of the free-text search,
// same "no re-fetch" pattern as everything else in this file. exchanges is a Set: empty
// means "no exchange restriction" (not "match nothing").
let analysisFilters = { dateFrom: "", dateTo: "", address: "", txid: "", exchanges: new Set() };

function toggleAnalysisFilterPanel() {
    const panel = document.getElementById("analysisFilterPanel");
    const isHidden = panel.style.display === "none";
    panel.style.display = isHidden ? "block" : "none";
    if (isHidden) renderFilterExchangeList();
}

// catalogData (populated on load and after every upload) already has one entry per
// imported file with its detected exchange - cheaper and more reliable than scanning
// whichever analysis tab happens to be cached.
function getAvailableExchanges() {
    const set = new Set();
    Object.values(catalogData).forEach(f => { if (f.exchange) set.add(f.exchange); });
    return Array.from(set).sort();
}

function renderFilterExchangeList() {
    const list = document.getElementById("filterExchangeList");
    const exchanges = getAvailableExchanges();
    if (!exchanges.length) {
        list.innerHTML = '<span style="font-size:12px;color:var(--text-dim);">No files imported yet.</span>';
        return;
    }
    list.innerHTML = exchanges.map(ex => `
        <label class="filter-exchange-item">
            <input type="checkbox" ${analysisFilters.exchanges.has(ex) ? "checked" : ""} onchange="onExchangeFilterChange('${ex}', this.checked)">
            ${escapeHtml(ex)}
        </label>
    `).join("");
}

function onExchangeFilterChange(exchange, checked) {
    if (checked) analysisFilters.exchanges.add(exchange);
    else analysisFilters.exchanges.delete(exchange);
    applyAnalysisFilterState();
}

// Custom dd/mm/yyyy calendar popup for the date-range filter fields, replacing the native
// <input type="date"> whose display format follows the browser locale (often mm/dd/yyyy).
// Each field is a hidden input (id="filterDateFrom"/"filterDateTo", ISO "YYYY-MM-DD" value,
// read directly by matchesFilters/onAnalysisFilterChange) paired with a readonly display
// input showing the formatted date. A single popup instance is reused for both fields.
let dpField = null, dpDisplayEl = null, dpViewYear = null, dpViewMonth = null;

function fmtDateDMY(iso) {
    if (!iso) return "";
    const [y, m, d] = iso.split("-");
    return `${d}/${m}/${y}`;
}

function setDateFilterValue(fieldId, iso) {
    document.getElementById(fieldId).value = iso || "";
    document.getElementById(fieldId + "Display").value = fmtDateDMY(iso);
}

function openDatePicker(fieldId, displayEl) {
    const popup = document.getElementById("datePickerPopup");
    if (dpField === fieldId && popup.style.display !== "none") {
        closeDatePicker();
        return;
    }
    dpField = fieldId;
    dpDisplayEl = displayEl;
    const iso = document.getElementById(fieldId).value;
    const base = iso ? new Date(iso + "T00:00:00") : new Date();
    dpViewYear = base.getFullYear();
    dpViewMonth = base.getMonth();

    const rect = displayEl.getBoundingClientRect();
    popup.style.display = "block";
    popup.style.top = `${rect.bottom + 4}px`;
    popup.style.left = `${Math.max(8, rect.left)}px`;
    dpRender();

    document.addEventListener("mousedown", dpOutsideClick, true);
    document.addEventListener("keydown", dpEscape, true);
}

function closeDatePicker() {
    document.getElementById("datePickerPopup").style.display = "none";
    dpField = null;
    dpDisplayEl = null;
    document.removeEventListener("mousedown", dpOutsideClick, true);
    document.removeEventListener("keydown", dpEscape, true);
}

function dpOutsideClick(e) {
    const popup = document.getElementById("datePickerPopup");
    if (popup.contains(e.target) || e.target === dpDisplayEl) return;
    closeDatePicker();
}

function dpEscape(e) {
    if (e.key === "Escape") closeDatePicker();
}

function dpChangeMonth(delta) {
    dpViewMonth += delta;
    if (dpViewMonth < 0) { dpViewMonth = 11; dpViewYear--; }
    else if (dpViewMonth > 11) { dpViewMonth = 0; dpViewYear++; }
    dpRender();
}

const DP_MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"];

function dpRender() {
    document.getElementById("dpTitle").textContent = `${DP_MONTH_NAMES[dpViewMonth]} ${dpViewYear}`;

    const selectedIso = dpField ? document.getElementById(dpField).value : "";
    const todayIso = new Date().toISOString().slice(0, 10);

    const firstOfMonth = new Date(dpViewYear, dpViewMonth, 1);
    const startOffset = (firstOfMonth.getDay() + 6) % 7; // Monday-first week
    const gridStart = new Date(dpViewYear, dpViewMonth, 1 - startOffset);

    let html = "";
    for (let i = 0; i < 42; i++) {
        const cell = new Date(gridStart.getFullYear(), gridStart.getMonth(), gridStart.getDate() + i);
        const cellIso = `${cell.getFullYear()}-${String(cell.getMonth() + 1).padStart(2, "0")}-${String(cell.getDate()).padStart(2, "0")}`;
        const outside = cell.getMonth() !== dpViewMonth;
        const cls = ["dp-day"];
        if (outside) cls.push("dp-outside");
        if (cellIso === todayIso) cls.push("dp-today");
        if (cellIso === selectedIso) cls.push("dp-selected");
        html += `<button type="button" class="${cls.join(" ")}" onclick="dpSelectDay('${cellIso}')">${cell.getDate()}</button>`;
    }
    document.getElementById("dpGrid").innerHTML = html;
}

function dpSelectDay(iso) {
    if (!dpField) return;
    setDateFilterValue(dpField, iso);
    closeDatePicker();
    onAnalysisFilterChange();
}

function dpToday() {
    if (!dpField) return;
    const iso = new Date().toISOString().slice(0, 10);
    setDateFilterValue(dpField, iso);
    closeDatePicker();
    onAnalysisFilterChange();
}

function dpClear() {
    if (!dpField) return;
    setDateFilterValue(dpField, "");
    closeDatePicker();
    onAnalysisFilterChange();
}

function onAnalysisFilterChange() {
    analysisFilters.dateFrom = document.getElementById("filterDateFrom").value;
    analysisFilters.dateTo = document.getElementById("filterDateTo").value;
    analysisFilters.address = document.getElementById("filterAddress").value.trim().toLowerCase();
    analysisFilters.txid = document.getElementById("filterTxid").value.trim().toLowerCase();
    applyAnalysisFilterState();
}

function activeFilterCount() {
    let n = 0;
    if (analysisFilters.dateFrom) n++;
    if (analysisFilters.dateTo) n++;
    if (analysisFilters.address) n++;
    if (analysisFilters.txid) n++;
    if (analysisFilters.exchanges.size) n++;
    return n;
}

function applyAnalysisFilterState() {
    const count = activeFilterCount();
    const btn = document.getElementById("analysisFilterToggleBtn");
    btn.textContent = count ? `🔎 Filters (${count}) ▾` : "🔎 Filters ▾";
    btn.classList.toggle("active", count > 0);
    reapplyAnalysisSearch();
}

function clearAnalysisFilters() {
    setDateFilterValue("filterDateFrom", "");
    setDateFilterValue("filterDateTo", "");
    document.getElementById("filterAddress").value = "";
    document.getElementById("filterTxid").value = "";
    analysisFilters = { dateFrom: "", dateTo: "", address: "", txid: "", exchanges: new Set() };
    renderFilterExchangeList();
    applyAnalysisFilterState();
}

// row: { date: ISO string or null, exchange, address, txid }
function matchesFilters(row) {
    const day = row.date ? row.date.slice(0, 10) : null;
    if (analysisFilters.dateFrom && (!day || day < analysisFilters.dateFrom)) return false;
    if (analysisFilters.dateTo && (!day || day > analysisFilters.dateTo)) return false;
    if (analysisFilters.exchanges.size && !analysisFilters.exchanges.has(row.exchange)) return false;
    if (analysisFilters.address && !(row.address || "").toLowerCase().includes(analysisFilters.address)) return false;
    if (analysisFilters.txid && !(row.txid || "").toLowerCase().includes(analysisFilters.txid)) return false;
    return true;
}

// Transfers tab: same person / different people filter, applied client-side on top of the
// search filter (no re-fetch needed - the choice doesn't change what the server already sent).
let transferPersonFilter = "all";

function personFilterChip(value, label) {
    const active = transferPersonFilter === value ? " active" : "";
    return `<button class="filter-chip${active}" onclick="setTransferPersonFilter('${value}')">${escapeHtml(label)}</button>`;
}

function setTransferPersonFilter(value) {
    transferPersonFilter = value;
    transfersPageSize = ANALYSIS_PAGE_SIZE;
    renderTransfers(document.getElementById("analysisContent"));
}

// Date/Amount sorting for Amounts and Transfers, applied client-side on top of whatever
// order the server returned (default order is preserved until the user actually clicks a
// column, same "no re-fetch needed" pattern as search/person filter).
let amountsSort = { key: null, dir: "desc" };
let transfersSort = { key: null, dir: "desc" };

// Amounts/Transfers can easily reach thousands of rows - only the first PAGE_SIZE are
// rendered, with a "Load more" button bumping the count. Reset to the first page whenever
// the search/filter/sort actually changes (see reapplyAnalysisSearch, setAmountsSort,
// setTransfersSort, setTransferPersonFilter) so a new query doesn't stay scrolled deep in.
const ANALYSIS_PAGE_SIZE = 50;
let amountsPageSize = ANALYSIS_PAGE_SIZE;
let transfersPageSize = ANALYSIS_PAGE_SIZE;

function loadMoreAmounts() {
    amountsPageSize += ANALYSIS_PAGE_SIZE;
    renderAmounts(document.getElementById("analysisContent"));
}

function loadMoreTransfers() {
    transfersPageSize += ANALYSIS_PAGE_SIZE;
    renderTransfers(document.getElementById("analysisContent"));
}

function loadMoreButtonHtml(shownCount, totalCount, onClickFnName) {
    if (shownCount >= totalCount) return "";
    const remaining = totalCount - shownCount;
    return `<div class="load-more-row">
        <button class="btn small" onclick="${onClickFnName}()">Load ${Math.min(remaining, ANALYSIS_PAGE_SIZE)} more (${remaining} remaining)</button>
    </div>`;
}

// value used to compare two rows for a given sort key - "amount" prefers the USD-equivalent
// when available (so mixed-currency lists still sort meaningfully), falling back to the raw
// amount. Missing values sort last regardless of direction.
function sortValue(row, key) {
    if (key === "date") return row.date ? new Date(row.date).getTime() : null;
    if (key === "amount") {
        const v = row.amount_usd !== null && row.amount_usd !== undefined ? row.amount_usd : row.amount;
        return v === null || v === undefined ? null : v;
    }
    return null;
}

function applySort(list, state, extractor) {
    if (!state.key) return list;
    const withIdx = list.map((item, i) => ({ item, i, v: extractor(item, state.key) }));
    withIdx.sort((a, b) => {
        if (a.v === null && b.v === null) return a.i - b.i;
        if (a.v === null) return 1;   // missing values always last
        if (b.v === null) return -1;
        if (a.v !== b.v) return state.dir === "asc" ? a.v - b.v : b.v - a.v;
        return a.i - b.i;  // stable tie-break
    });
    return withIdx.map(x => x.item);
}

function sortableTh(label, key, state, setterName) {
    const active = state.key === key;
    const arrow = active ? (state.dir === "asc" ? " ▲" : " ▼") : "";
    return `<th class="sortable-th" onclick="${setterName}('${key}')">${escapeHtml(label)}${arrow}</th>`;
}

function sortChip(label, key, state, setterName) {
    const active = state.key === key;
    const arrow = active ? (state.dir === "asc" ? " ▲" : " ▼") : "";
    return `<button class="filter-chip${active ? " active" : ""}" onclick="${setterName}('${key}')">${escapeHtml(label)}${arrow}</button>`;
}

// Matched transfer "cards" have no top-level date/amount (they're a withdrawal+deposit
// pair) - sort by the withdrawal side, which is when the money actually moved.
function transferSortValue(p, key) { return sortValue(p.withdrawal, key); }

function toggleSort(state, key) {
    if (state.key === key) { state.dir = state.dir === "asc" ? "desc" : "asc"; }
    else { state.key = key; state.dir = "desc"; }
}

function setAmountsSort(key) {
    toggleSort(amountsSort, key);
    amountsPageSize = ANALYSIS_PAGE_SIZE;
    renderAmounts(document.getElementById("analysisContent"));
}

function setTransfersSort(key) {
    toggleSort(transfersSort, key);
    transfersPageSize = ANALYSIS_PAGE_SIZE;
    renderTransfers(document.getElementById("analysisContent"));
}

function matchesSearch(text) {
    if (!analysisSearchQuery) return true;
    return (text || "").toLowerCase().includes(analysisSearchQuery);
}

function highlightMatch(text) {
    const safe = escapeHtml(text == null ? "" : String(text));
    if (!analysisSearchQuery) return safe;
    const idx = safe.toLowerCase().indexOf(escapeHtml(analysisSearchQuery).toLowerCase());
    if (idx === -1) return safe;
    return safe.slice(0, idx) + '<mark class="search-hit">' + safe.slice(idx, idx + analysisSearchQuery.length) + "</mark>" + safe.slice(idx + analysisSearchQuery.length);
}

function onAnalysisSearch(value) {
    analysisSearchQuery = value.trim().toLowerCase();
    document.getElementById("analysisSearchClear").style.display = analysisSearchQuery ? "inline-block" : "none";
    reapplyAnalysisSearch();
}

function clearAnalysisSearch() {
    document.getElementById("analysisSearch").value = "";
    onAnalysisSearch("");
}

// Re-renders the current tab from already-fetched data (no network round-trip per
// keystroke). Falls back to a full load if nothing's cached yet for that tab.
function reapplyAnalysisSearch() {
    amountsPageSize = ANALYSIS_PAGE_SIZE;
    transfersPageSize = ANALYSIS_PAGE_SIZE;
    const container = document.getElementById("analysisContent");
    if (currentAnalysisTab === "addresses") { if (cachedAddresses) renderAddresses(container); else loadAddresses(container); }
    else if (currentAnalysisTab === "amounts") { if (cachedAmounts) renderAmounts(container); else loadAmounts(container); }
    else if (currentAnalysisTab === "transfers") { if (cachedTransfers) renderTransfers(container); else loadTransfers(container); }
    else if (currentAnalysisTab === "chains") { if (cachedChains) renderChains(container); else loadChains(container); }
    else if (currentAnalysisTab === "graph") focusGraphSearchMatch();
}

function copyAddress(address) {
    navigator.clipboard.writeText(address).then(
        () => showToast("Address copied", false),
        () => showToast("Couldn't copy address", true)
    );
}

function hideWallet(address) {
    fetch("/addresses/exclude", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address })
    }).then(() => {
        showToast("Wallet hidden - restore it anytime from the sidebar", false);
        loadAnalysisTab(currentAnalysisTab);
        refreshHiddenWalletsSidebar();
    });
}

function restoreWallet(address) {
    fetch("/addresses/include", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address })
    }).then(() => {
        loadAnalysisTab(currentAnalysisTab);
        refreshHiddenWalletsSidebar();
    });
}

function refreshHiddenWalletsSidebar() {
    const sidebar = document.getElementById("hiddenWalletsSidebar");
    fetch("/addresses/excluded").then(r => r.json()).then(list => {
        if (!list.length) { sidebar.style.display = "none"; sidebar.innerHTML = ""; return; }
        sidebar.style.display = "block";
        sidebar.innerHTML = `<h3>Hidden wallets (${list.length})</h3>` + list.map(item => `
            <div class="hidden-wallet-item">
                <div class="hidden-wallet-addr">${addressCellHtml(item.address, false)}</div>
                <div class="hidden-wallet-meta">${item.occurrence_count} occurrence(s)${item.suspect_names.length ? " · " + escapeHtml(item.suspect_names.join(", ")) : ""}</div>
                ${item.note ? `<div class="hidden-wallet-meta" style="color:var(--text-dim);">📝 ${escapeHtml(item.note)}</div>` : ""}
                <button class="btn small" onclick="restoreWallet('${escapeHtml(item.address).replace(/'/g, "\\'")}')">Restore</button>
            </div>
        `).join("");
    }).catch(() => {});
}

function loadAmounts(container) {
    fetch("/analysis/amounts" + suspectsQueryParam()).then(r => r.json()).then(data => {
        cachedAmounts = data;
        amountsPageSize = ANALYSIS_PAGE_SIZE;
        renderAmounts(container);
    }).catch(err => {
        container.innerHTML = `<div class="analysis-empty">Error loading data: ${escapeHtml(String(err))}</div>`;
    });
}

function renderAmounts(container) {
    const data = cachedAmounts || [];
    if (!data.length) {
        container.innerHTML = '<div class="analysis-empty">No transactions found yet - import some files first.</div>';
        return;
    }
    const searchFiltered = data.filter(r =>
        (matchesSearch(r.external_address) || matchesSearch(r.txid)) &&
        matchesFilters({ date: r.date, exchange: r.exchange, address: r.external_address, txid: r.txid })
    );
    if (!searchFiltered.length) {
        container.innerHTML = `<div class="analysis-empty">No transaction matches the current search/filters.</div>`;
        return;
    }
    const filtered = applySort(searchFiltered, amountsSort, sortValue);
    const sortNote = amountsSort.key
        ? `sorted by ${amountsSort.key} (${amountsSort.dir === "asc" ? "ascending" : "descending"})`
        : "sorted largest to smallest (USD-equivalent first when available)";
    const page = filtered.slice(0, amountsPageSize);
    let html = `<div class="results-note">${page.length} of ${filtered.length} matching transaction(s) shown (${data.length} total), ${sortNote}. Click a column header to sort.</div>`;
    html += `<div class="table-wrap"><table><thead><tr><th>Suspect</th><th>Exchange</th><th>Type</th>${sortableTh("Amount", "amount", amountsSort, "setAmountsSort")}${sortableTh("Date", "date", amountsSort, "setAmountsSort")}<th>Address flow</th><th>TXID</th></tr></thead><tbody>`;
    page.forEach(r => {
        // Money's real-world direction flips the FROM/TO meaning of the same two columns:
        // a deposit's external address is the sender (FROM) and the exchange's own address is
        // the receiver (TO); a withdrawal is the mirror image.
        const isDeposit = r.file_type === "deposit";
        const addrLabel = isDeposit ? "FROM" : "TO";
        const exAddrLabel = isDeposit ? "TO" : "FROM";
        html += `<tr>
            <td>${escapeHtml(r.suspect_name)}</td>
            <td>${escapeHtml(r.exchange)}</td>
            <td><span class="badge ${r.file_type}">${TYPE_LABELS[r.file_type] || r.file_type}</span></td>
            <td>${fmtAmount(r.amount, r.currency)}${r.amount_usd ? ' <span style="color:var(--text-dim);font-size:11.5px;">(' + fmtUsd(r.amount_usd) + ')</span>' : ""}</td>
            <td>${fmtDate(r.date)}</td>
            <td>
                <div class="flow-line"><span class="flow-tag">${addrLabel}</span> ${truncMono(r.external_address)} ${categoryBadgeHtml(r.external_address, r.category, r.category_color)}${noteBadgeHtml(r.external_address, r.note)}</div>
                <div class="flow-arrow">⇄</div>
                <div class="flow-line"><span class="flow-tag">${exAddrLabel}</span> ${r.exchange_address ? truncMono(r.exchange_address) + " " + categoryBadgeHtml(r.exchange_address, r.exchange_address_category, r.exchange_address_category_color) + noteBadgeHtml(r.exchange_address, r.exchange_address_note) : '<span style="color:var(--text-dim);">-</span>'}</div>
            </td>
            <td>${truncMono(r.txid)}</td>
        </tr>`;
    });
    html += "</tbody></table></div>";
    html += loadMoreButtonHtml(page.length, filtered.length, "loadMoreAmounts");
    container.innerHTML = html;
}

function loadTransfers(container) {
    fetch("/analysis/transfers" + suspectsQueryParam()).then(r => r.json()).then(data => {
        cachedTransfers = data;
        transfersPageSize = ANALYSIS_PAGE_SIZE;
        renderTransfers(container);
    }).catch(err => {
        container.innerHTML = `<div class="analysis-empty">Error loading data: ${escapeHtml(String(err))}</div>`;
    });
}

// Renders one withdrawal<->deposit hop card. Shared by the Transfers tab (one hop per card)
// and the Chains tab (several of these stacked per chain, connected by transferChainLinkHtml).
function transferCardHtml(p) {
    const w = p.withdrawal, d = p.deposit;
    const contextAddress = p.address || w.external_address || d.external_address;
    const who = p.same_suspect
        ? `<b>${escapeHtml(w.suspect_name)}</b> sent to themself`
        : `<b>${escapeHtml(w.suspect_name)}</b> sent to <b>${escapeHtml(d.suspect_name)}</b>`;
    const viaLabel = p.match_type === "txid"
        ? `matching TX hash <span class="addr-mono">${highlightMatch(w.txid)}</span>`
        : `shared address <span class="addr-mono">${highlightMatch(p.address)}</span>`;
    return `<div class="transfer-card">
        <div class="addr-line">
            ${p.match_type === "txid" ? '<span class="same-person-badge" style="background:rgba(62,207,142,0.18);color:var(--success);">✓ CONFIRMED (SAME TXID)</span> ' : '<span class="unverified-badge" style="background:rgba(139,147,167,0.18);color:var(--text-dim);">PROBABLE (SAME ADDRESS)</span> '}
            ${who} — via ${viaLabel}
            ${p.same_suspect ? '<span class="same-person-badge">SAME PERSON</span>' : '<span class="diff-suspect-badge">DIFFERENT PEOPLE</span>'}
            ${!p.same_exchange ? '<span class="same-person-badge" style="background:rgba(139,147,167,0.18);color:var(--text-dim);">CROSS-EXCHANGE</span>' : ""}
            ${!p.amount_matched ? '<span class="unverified-badge" title="Amount and/or currency could not be confirmed between the two sides - verify manually.">⚠ AMOUNT UNVERIFIED</span>' : ""}
        </div>
        <div class="transfer-flow">
            <div class="transfer-side">
                <div class="label">Sent (withdrawal)</div>
                <div class="exchange-name">${escapeHtml(w.exchange)}</div>
                <div>${escapeHtml(w.suspect_name)}</div>
                <div class="amount">${fmtAmount(w.amount, w.currency)}</div>
                <div style="color:var(--text-dim);font-size:12px;">${fmtDate(w.date)}</div>
                <div class="txid-line">TX Hash: <span class="addr-mono">${highlightMatch(w.txid || "-")}</span></div>
                <div class="txid-line">Address: <span class="addr-mono">${highlightMatch(w.external_address || "-")}</span> ${categoryBadgeHtml(w.external_address, w.category, w.category_color)}${noteBadgeHtml(w.external_address, w.note)}</div>
                ${w.exchange_address ? `<div class="txid-line">Exchange address: <span class="addr-mono">${highlightMatch(w.exchange_address)}</span> ${categoryBadgeHtml(w.exchange_address, w.exchange_address_category, w.exchange_address_category_color)}${noteBadgeHtml(w.exchange_address, w.exchange_address_note)}</div>` : ""}
            </div>
            <div class="transfer-arrow">→</div>
            <div class="transfer-side">
                <div class="label">Received (deposit)</div>
                <div class="exchange-name">${escapeHtml(d.exchange)}</div>
                <div>${escapeHtml(d.suspect_name)}</div>
                <div class="amount">${fmtAmount(d.amount, d.currency)}</div>
                <div style="color:var(--text-dim);font-size:12px;">${fmtDate(d.date)}</div>
                <div class="txid-line">TX Hash: <span class="addr-mono">${highlightMatch(d.txid || "-")}</span></div>
                <div class="txid-line">Address: <span class="addr-mono">${highlightMatch(d.external_address || "-")}</span> ${categoryBadgeHtml(d.external_address, d.category, d.category_color)}${noteBadgeHtml(d.external_address, d.note)}</div>
                ${d.exchange_address ? `<div class="txid-line">Exchange address: <span class="addr-mono">${highlightMatch(d.exchange_address)}</span> ${categoryBadgeHtml(d.exchange_address, d.exchange_address_category, d.exchange_address_category_color)}${noteBadgeHtml(d.exchange_address, d.exchange_address_note)}</div>` : ""}
            </div>
        </div>
        <div class="transfer-gap">
            Time gap: ${p.gap_hours} hour(s)
            ${contextAddress ? `<button class="btn small" style="margin-left:10px;" onclick="viewWalletHistory('${escapeHtml(contextAddress).replace(/'/g, "\\'")}')">🔍 View this wallet's history</button>` : ""}
        </div>
    </div>`;
}

// Jumps to Amounts, pre-filtered to just this wallet address (clearing any other filter/
// search) so the transactions right before and after this transfer are easy to see for
// context - what else moved through this same wallet, on either side of it.
function viewWalletHistory(address) {
    if (!address) return;
    analysisSearchQuery = "";
    document.getElementById("analysisSearch").value = "";
    document.getElementById("analysisSearchClear").style.display = "none";

    analysisFilters = { dateFrom: "", dateTo: "", address: address.trim().toLowerCase(), txid: "", exchanges: new Set() };
    setDateFilterValue("filterDateFrom", "");
    setDateFilterValue("filterDateTo", "");
    document.getElementById("filterTxid").value = "";
    document.getElementById("filterAddress").value = address;
    document.getElementById("analysisFilterPanel").style.display = "block";
    renderFilterExchangeList();
    const filterBtn = document.getElementById("analysisFilterToggleBtn");
    filterBtn.textContent = "🔎 Filters (1) ▾";
    filterBtn.classList.add("active");

    switchAnalysisTab("amounts");
}

function loadChains(container) {
    fetch("/analysis/chains" + suspectsQueryParam()).then(r => r.json()).then(data => {
        cachedChains = data;
        renderChains(container);
    }).catch(err => {
        container.innerHTML = `<div class="analysis-empty">Error loading data: ${escapeHtml(String(err))}</div>`;
    });
}

function renderChains(container) {
    const allChains = cachedChains || [];
    if (!allChains.length) {
        container.innerHTML = '<div class="analysis-empty">No multi-hop chains detected yet - chains need at least 2 linked transfers (someone receiving money, then sending it on).</div>';
        return;
    }

    const chains = allChains.filter(c => c.some(p =>
        (matchesSearch(p.address) || matchesSearch(p.withdrawal.txid) || matchesSearch(p.deposit.txid)) &&
        (matchesFilters({ date: p.withdrawal.date, exchange: p.withdrawal.exchange, address: p.address, txid: p.withdrawal.txid }) ||
         matchesFilters({ date: p.deposit.date, exchange: p.deposit.exchange, address: p.address, txid: p.deposit.txid }))
    ));
    if (!chains.length) {
        container.innerHTML = `<div class="analysis-empty">No chain matches the current search/filters.</div>`;
        return;
    }

    let html = `<div class="results-note">
        ${chains.length}${analysisSearchQuery ? ` of ${allChains.length}` : ""} chain(s) shown — each links 2+ already-detected transfers where
        the same person received money in one, then sent it on in another shortly after. Longest first.
    </div>`;

    chains.forEach(chain => {
        const start = chain[0].withdrawal, end = chain[chain.length - 1].deposit;
        const peopleInChain = new Set();
        chain.forEach(p => { peopleInChain.add(p.withdrawal.suspect_id); peopleInChain.add(p.deposit.suspect_id); });
        const totalHours = Math.round((new Date(end.date) - new Date(start.date)) / 36000) / 100;

        html += `<div class="chain-card">
            <div class="chain-header">
                <span class="chain-path">${chain.map(p => `<b>${escapeHtml(p.withdrawal.suspect_name)}</b>`).join(" → ")} → <b>${escapeHtml(end.suspect_name)}</b></span>
                <span class="same-person-badge">${chain.length} HOPS</span>
                <span class="${peopleInChain.size > 1 ? 'diff-suspect-badge' : 'same-person-badge'}">${peopleInChain.size} ${peopleInChain.size > 1 ? 'PEOPLE' : 'PERSON'}</span>
                <span class="unverified-badge" style="background:rgba(139,147,167,0.18);color:var(--text-dim);">~${totalHours}h START TO END</span>
            </div>`;
        chain.forEach((hop, i) => {
            html += transferCardHtml(hop);
            if (i < chain.length - 1) {
                const nextHop = chain[i + 1];
                const gapToNext = Math.round((new Date(nextHop.withdrawal.date) - new Date(hop.deposit.date)) / 36000) / 100;
                html += `<div class="chain-link">↓ <b>${escapeHtml(hop.deposit.suspect_name)}</b> then sent this on, ${gapToNext} hour(s) later ↓</div>`;
            }
        });
        html += `</div>`;
    });

    container.innerHTML = html;
}

function renderTransfers(container) {
    const allMatched = (cachedTransfers && cachedTransfers.matched) || [];
    const allUnmatched = (cachedTransfers && cachedTransfers.unmatched) || [];
    if (!allMatched.length && !allUnmatched.length) {
        container.innerHTML = '<div class="analysis-empty">No wallet-to-wallet transfers detected yet.</div>';
        return;
    }

    // A pair matches the filter panel if either leg (withdrawal or deposit) falls within the
    // chosen date range/exchange - the two sides can land on different dates/exchanges, and
    // hiding the whole transfer because one side misses would be more confusing than useful.
    const searchMatched = allMatched.filter(p =>
        (matchesSearch(p.address) || matchesSearch(p.withdrawal.txid) || matchesSearch(p.deposit.txid)) &&
        (matchesFilters({ date: p.withdrawal.date, exchange: p.withdrawal.exchange, address: p.address, txid: p.withdrawal.txid }) ||
         matchesFilters({ date: p.deposit.date, exchange: p.deposit.exchange, address: p.address, txid: p.deposit.txid }))
    );
    const unmatched = allUnmatched.filter(u =>
        (matchesSearch(u.address) || matchesSearch(u.txid)) &&
        matchesFilters({ date: u.date, exchange: u.exchange, address: u.address, txid: u.txid })
    );
    const diffPeopleCount = searchMatched.filter(p => !p.same_suspect).length;
    const samePersonCount = searchMatched.length - diffPeopleCount;

    const matched = applySort(searchMatched.filter(p =>
        transferPersonFilter === "all" ||
        (transferPersonFilter === "same" && p.same_suspect) ||
        (transferPersonFilter === "different" && !p.same_suspect)
    ), transfersSort, transferSortValue);
    const unmatchedSorted = applySort(unmatched, transfersSort, sortValue);

    let html = `<div class="person-filter-row">
        <span class="filter-label">Show:</span>
        ${personFilterChip("all", `All (${searchMatched.length})`)}
        ${personFilterChip("same", `Same person (${samePersonCount})`)}
        ${personFilterChip("different", `Different people (${diffPeopleCount})`)}
        <span class="filter-label" style="margin-left:10px;">Sort by:</span>
        ${sortChip("Date", "date", transfersSort, "setTransfersSort")}
        ${sortChip("Amount", "amount", transfersSort, "setTransfersSort")}
    </div>`;

    if (!matched.length && !unmatched.length) {
        html += analysisSearchQuery
            ? `<div class="analysis-empty">No address or TXID matches "${escapeHtml(analysisSearchQuery)}".</div>`
            : '<div class="analysis-empty">No transfer matches this filter.</div>';
        container.innerHTML = html;
        return;
    }

    const txidCount = matched.filter(p => p.match_type === "txid").length;
    const matchedPage = matched.slice(0, transfersPageSize);

    html += `<div class="results-note">
        ${matchedPage.length} of ${matched.length}${(analysisSearchQuery || transferPersonFilter !== "all") ? ` (${allMatched.length} total)` : ""} confirmed transfer(s) shown — <b>${txidCount}</b> proven by matching <b>TXID</b> (same blockchain tx on both sides),
        ${matched.length - txidCount} inferred from a shared address.
        ${unmatched.length ? `${unmatched.length} movement(s) could not be confidently paired — see below.` : ""}
    </div>`;

    matchedPage.forEach(p => { html += transferCardHtml(p); });
    html += loadMoreButtonHtml(matchedPage.length, matched.length, "loadMoreTransfers");

    if (unmatched.length) {
        html += `<div class="panel-title" style="margin-top:24px;">Unmatched movements (needs manual review)</div>`;
        html += `<div class="table-wrap"><table><thead><tr><th>Direction</th><th>Suspect</th><th>Exchange</th>${sortableTh("Amount", "amount", transfersSort, "setTransfersSort")}${sortableTh("Date", "date", transfersSort, "setTransfersSort")}<th>Address</th><th>Exchange address</th><th>TXID</th></tr></thead><tbody>`;
        unmatchedSorted.forEach(u => {
            html += `<tr>
                <td><span class="badge ${u.direction}">${TYPE_LABELS[u.direction] || u.direction}</span></td>
                <td>${escapeHtml(u.suspect_name)}</td>
                <td>${escapeHtml(u.exchange)}</td>
                <td>${fmtAmount(u.amount, u.currency)}</td>
                <td>${fmtDate(u.date)}</td>
                <td>${truncMono(u.address)} ${categoryBadgeHtml(u.address, u.category, u.category_color)}${noteBadgeHtml(u.address, u.note)}</td>
                <td>${u.exchange_address ? truncMono(u.exchange_address) + " " + categoryBadgeHtml(u.exchange_address, u.exchange_address_category, u.exchange_address_category_color) + noteBadgeHtml(u.exchange_address, u.exchange_address_note) : '<span style="color:var(--text-dim);">-</span>'}</td>
                <td>${truncMono(u.txid)}</td>
            </tr>`;
        });
        html += "</tbody></table></div>";
    }

    container.innerHTML = html;
}

let graphNetworkInstance = null;

let graphToolbarHideTimer = null;

let cachedGraphData = null;

let graphColorFilters = { cross: true, same: true };

// Node drag positions, persisted in localStorage (keyed by node id, which embeds the
// suspect's UUID or the address so it's stable across reloads and Save/Load case) so
// rearranging the graph isn't lost on a tab switch or a page refresh.
function loadGraphNodePositions() {
    try { return JSON.parse(localStorage.getItem("cryptolink-graph-positions") || "{}"); }
    catch { return {}; }
}
function saveGraphNodePositions() {
    localStorage.setItem("cryptolink-graph-positions", JSON.stringify(graphNodePositions));
}
let graphNodePositions = loadGraphNodePositions();

function loadGraph(container) {
    container.innerHTML = `
        <div class="graph-controls">
            <div class="results-note" style="margin-bottom:0;">Only accounts linked by a shared wallet or a confirmed TXID transfer are shown. Arrows show the direction money moved. Hover a wallet node for actions, click an arrow to see the transactions behind it.</div>
            <div class="graph-legend">
                <span class="legend-item"><span class="legend-dot" style="background:#4f8cff;"></span> Suspect</span>
                <label class="graph-toggle"><input type="checkbox" id="graphFilterCross" ${graphColorFilters.cross ? "checked" : ""} onchange="setGraphColorFilter('cross', this.checked)"> <span class="legend-dot" style="background:#f0556b;"></span> Wallet shared between different people</label>
                <label class="graph-toggle"><input type="checkbox" id="graphFilterSame" ${graphColorFilters.same ? "checked" : ""} onchange="setGraphColorFilter('same', this.checked)"> <span class="legend-dot" style="background:#f5a623;"></span> Wallet shared by the same person</label>
                <span class="legend-item"><span class="legend-dot" style="background:#3ecf8e;"></span> Confirmed TXID transfer (direct link)</span>
            </div>
        </div>
        <div class="graph-canvas-wrap">
            <div id="graphCanvas"></div>
            <div id="graphNodeToolbar" class="node-toolbar">
                <span class="icon-btn" id="graphNodeNoteIcon" style="display:none;cursor:help;">📝</span>
                <button class="icon-btn" title="Copy address" onclick="copyAddress(document.getElementById('graphNodeToolbar').dataset.address)">📋</button>
                <button class="icon-btn danger" title="Hide this wallet" onclick="hideWallet(document.getElementById('graphNodeToolbar').dataset.address)">🗑️</button>
            </div>
        </div>
        <div id="graphEdgeDetail" class="graph-edge-detail" style="display:none;"></div>
    `;

    const toolbar = document.getElementById("graphNodeToolbar");
    toolbar.addEventListener("mouseenter", () => clearTimeout(graphToolbarHideTimer));
    toolbar.addEventListener("mouseleave", scheduleHideGraphToolbar);

    fetch("/analysis/graph" + suspectsQueryParam()).then(r => r.json()).then(data => {
        cachedGraphData = data;
        renderGraph();
    }).catch(err => {
        const canvas = document.getElementById("graphCanvas");
        if (canvas) canvas.innerHTML = `<div class="analysis-empty">Error loading graph: ${escapeHtml(String(err))}</div>`;
    });
}

// Toggling a legend checkbox re-filters the already-fetched graph in place, no re-fetch
// needed - same pattern as the Transfers same/different-person filter.
function setGraphColorFilter(kind, checked) {
    graphColorFilters[kind] = checked;
    renderGraph();
}

function renderGraph() {
    const data = cachedGraphData;
    const canvas = document.getElementById("graphCanvas");
    if (!data || !canvas) return;

    if (graphNetworkInstance) { graphNetworkInstance.destroy(); graphNetworkInstance = null; }
    hideGraphEdgeDetail();

    if (!data.nodes.length) {
        canvas.innerHTML = '<div class="analysis-empty">No wallet connects two different accounts yet.</div>';
        return;
    }

    const hiddenGroups = new Set();
    if (!graphColorFilters.cross) hiddenGroups.add("address_shared_cross_suspect");
    if (!graphColorFilters.same) hiddenGroups.add("address_shared_same_suspect");

    const keptNodeIds = new Set(data.nodes.filter(n => !hiddenGroups.has(n.group)).map(n => n.id));
    let visibleEdges = data.edges.filter(e => keptNodeIds.has(e.from) && keptNodeIds.has(e.to));

    // A suspect left with zero remaining edges once a wallet color is hidden would just be
    // an isolated dot - drop those too for a cleaner result.
    const connectedIds = new Set();
    visibleEdges.forEach(e => { connectedIds.add(e.from); connectedIds.add(e.to); });
    const visibleNodes = data.nodes.filter(n => {
        if (!keptNodeIds.has(n.id)) return false;
        return n.group !== "suspect" || connectedIds.has(n.id);
    });
    const finalIds = new Set(visibleNodes.map(n => n.id));
    visibleEdges = visibleEdges.filter(e => finalIds.has(e.from) && finalIds.has(e.to));

    if (!visibleNodes.length) {
        canvas.innerHTML = '<div class="analysis-empty">No wallet matches the current filters.</div>';
        return;
    }
    canvas.innerHTML = "";

    // Apply any positions saved from a previous drag (this session or a past one) and pin
    // them so the initial physics pass doesn't shuffle them - new nodes with no saved
    // position are left for physics to place normally, next to their now-fixed neighbors.
    visibleNodes.forEach(n => {
        const pos = graphNodePositions[n.id];
        if (pos) { n.x = pos.x; n.y = pos.y; n.fixed = { x: true, y: true }; }
    });

    const nodes = new vis.DataSet(visibleNodes);
    const edges = new vis.DataSet(visibleEdges);
    const isLightTheme = document.documentElement.getAttribute("data-theme") === "light";
    const toolbar = document.getElementById("graphNodeToolbar");

    const options = {
        nodes: {
            shape: "dot", size: 16, font: { color: isLightTheme ? "#1a1f2b" : "#e6e9f0", size: 12 },
            borderWidth: 2,
        },
        edges: {
            color: { color: "#3a4258", highlight: "#4f8cff" },
            smooth: { type: "continuous" },
            scaling: { min: 1, max: 8 },
        },
        groups: {
            suspect: { color: { background: "#4f8cff", border: "#2a4a8a" }, shape: "dot", size: 22 },
            address_shared_cross_suspect: { color: { background: "#f0556b", border: "#8a2030" } },
            address_shared_same_suspect: { color: { background: "#f5a623", border: "#8a5c10" } },
        },
        physics: { stabilization: true, barnesHut: { gravitationalConstant: -3000, springLength: 120 } },
        interaction: { hover: true, tooltipDelay: 100, dragNodes: true },
    };

    graphNetworkInstance = new vis.Network(canvas, { nodes, edges }, options);
    // Physics only runs for the initial layout. Once it settles, turn it off so dragging
    // one node repositions just that node instead of the whole graph reacting/reshuffling.
    graphNetworkInstance.once("stabilizationIterationsDone", () => {
        graphNetworkInstance.setOptions({ physics: false });
        // Capture where everything landed (including brand-new nodes physics just placed)
        // so the next render - another tab and back, or a page refresh - starts from here
        // instead of re-running physics from scratch.
        const allPositions = graphNetworkInstance.getPositions();
        Object.assign(graphNodePositions, allPositions);
        saveGraphNodePositions();
        // Release nodes that were pinned to a saved position - they should stay draggable,
        // not frozen in place now that physics is off anyway.
        nodes.get().forEach(n => {
            if (n.fixed) nodes.update({ id: n.id, fixed: { x: false, y: false } });
        });
    });

    // Persist wherever a node ends up after a manual drag, so the layout survives a tab
    // switch or a full page refresh instead of resetting to a fresh physics layout.
    graphNetworkInstance.on("dragEnd", params => {
        if (!params.nodes.length) return;
        const positions = graphNetworkInstance.getPositions(params.nodes);
        Object.assign(graphNodePositions, positions);
        saveGraphNodePositions();
    });

    // Click a suspect<->wallet or confirmed-transfer edge to see the actual transactions
    // behind it, listed below the graph.
    graphNetworkInstance.on("click", params => {
        if (params.nodes.length) return;
        if (params.edges.length) showGraphEdgeDetail(edges.get(params.edges[0]));
        else hideGraphEdgeDetail();
    });

    // Hover a wallet node -> show a small floating copy/hide toolbar next to it (suspect
    // nodes can't be hidden, so they get no toolbar). A short delay on hide lets the
    // mouse travel from the node onto the toolbar itself without it disappearing first.
    graphNetworkInstance.on("hoverNode", params => {
        const node = nodes.get(params.node);
        if (!node.group || !node.group.startsWith("address_shared")) return;
        clearTimeout(graphToolbarHideTimer);
        const pos = graphNetworkInstance.canvasToDOM(graphNetworkInstance.getPositions([params.node])[params.node]);
        toolbar.style.left = (pos.x + 16) + "px";
        toolbar.style.top = (pos.y - 14) + "px";
        toolbar.style.display = "flex";
        toolbar.dataset.address = node.title;
        const noteIcon = document.getElementById("graphNodeNoteIcon");
        noteIcon.style.display = node.note ? "inline-block" : "none";
        noteIcon.title = node.note || "";
    });
    graphNetworkInstance.on("blurNode", scheduleHideGraphToolbar);
    graphNetworkInstance.on("dragStart", () => { toolbar.style.display = "none"; });
    graphNetworkInstance.on("zoom", () => { toolbar.style.display = "none"; });

    focusGraphSearchMatch();
}

function scheduleHideGraphToolbar() {
    clearTimeout(graphToolbarHideTimer);
    graphToolbarHideTimer = setTimeout(() => {
        const toolbar = document.getElementById("graphNodeToolbar");
        if (toolbar) toolbar.style.display = "none";
    }, 250);
}

// Shows the transactions behind a clicked edge in a panel below the graph - a confirmed
// TXID transfer reuses the same transfer-card markup as the Transfers tab, a suspect<->wallet
// edge gets a small table of every transaction aggregated into that edge.
function showGraphEdgeDetail(edge) {
    const panel = document.getElementById("graphEdgeDetail");
    if (!panel) return;

    if (edge.transfer) {
        panel.innerHTML = `<div class="panel-title">Confirmed transfer</div>${transferCardHtml(edge.transfer)}`;
    } else if (edge.transactions) {
        const rows = (edge.transactions || []).slice().sort((a, b) => (a.date || "").localeCompare(b.date || ""));
        panel.innerHTML = `
            <div class="panel-title">${rows.length} transaction(s) between ${escapeHtml(edge.suspect_name)} and ${truncMono(edge.address)}</div>
            <div class="table-wrap"><table><thead><tr><th>Type</th><th>Exchange</th><th>Amount</th><th>Date</th><th>TXID</th></tr></thead><tbody>
                ${rows.map(o => `<tr>
                    <td><span class="badge ${o.file_type}">${TYPE_LABELS[o.file_type] || o.file_type}</span></td>
                    <td>${escapeHtml(o.exchange)}</td>
                    <td>${fmtAmount(o.amount, o.currency)}${o.amount_usd ? ' <span style="color:var(--text-dim);font-size:11px;">(' + fmtUsd(o.amount_usd) + ')</span>' : ""}</td>
                    <td>${fmtDate(o.date)}</td>
                    <td class="addr-mono">${highlightMatch(o.txid || "-")}</td>
                </tr>`).join("")}
            </tbody></table></div>
        `;
    } else {
        return;
    }
    panel.style.display = "block";
}

function hideGraphEdgeDetail() {
    const panel = document.getElementById("graphEdgeDetail");
    if (panel) panel.style.display = "none";
}

// Graph tab has no per-row list to filter, so the shared search box instead selects and
// centers on the matching wallet node (by address) or, failing that, the matching
// TXID-confirmed edge - same search box, same query, adapted to a graph instead of a table.
function focusGraphSearchMatch() {
    if (!graphNetworkInstance) return;
    if (!analysisSearchQuery) { graphNetworkInstance.unselectAll(); return; }

    const nodeMatch = graphNetworkInstance.body.data.nodes.get().find(n => matchesSearch(n.title));
    if (nodeMatch) {
        graphNetworkInstance.selectNodes([nodeMatch.id]);
        graphNetworkInstance.focus(nodeMatch.id, { scale: 1.3, animation: true });
        return;
    }
    const edgeMatch = graphNetworkInstance.body.data.edges.get().find(e => matchesSearch(e.title));
    if (edgeMatch) {
        graphNetworkInstance.selectEdges([edgeMatch.id]);
        graphNetworkInstance.focus(edgeMatch.from, { scale: 1.2, animation: true });
        return;
    }
    graphNetworkInstance.unselectAll();
}


// ---------------------------------------------------------------------------
// EXPORT SUSPECT FILTER
// ---------------------------------------------------------------------------

let suspectFilterSelected = new Set();

function syncSuspectFilterWithSuspects() {
    // keep filter selection in sync when suspects are added/removed -
    // new suspects default to "included", removed ones are dropped
    const currentIds = new Set(Object.keys(suspects));
    for (const id of currentIds) {
        if (!suspectFilterSelected.has(id)) suspectFilterSelected.add(id);
    }
    for (const id of Array.from(suspectFilterSelected)) {
        if (!currentIds.has(id)) suspectFilterSelected.delete(id);
    }
}

function toggleSuspectFilterPanel() {
    const panel = document.getElementById("suspectFilterPanel");
    const isHidden = panel.style.display === "none";
    panel.style.display = isHidden ? "block" : "none";
    if (isHidden) renderSuspectFilterList();
}

function toggleKnownWalletsPanel() {
    const panel = document.getElementById("knownWalletsPanel");
    panel.style.display = panel.style.display === "none" ? "block" : "none";
}

function commitKnownWallets() {
    const case_label = document.getElementById("knownWalletsLabel").value.trim();
    fetch("/known_wallets/commit", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ case_label })
    }).then(r => r.json()).then(data => {
        if (data.error) { showToast(data.error, true); return; }
        document.getElementById("knownWalletsLabel").value = "";
        document.getElementById("knownWalletsPanel").style.display = "none";
        showToast(`Added ${data.wallet_count} wallet(s) to the database as "${data.case_label}"`, false);
        if (currentMainTab === "analysis" && currentAnalysisTab === "addresses") loadAddresses(document.getElementById("analysisContent"));
    }).catch(err => showToast("Network error: " + err, true));
}

function addManualWalletCategory() {
    const address = document.getElementById("manualWalletAddress").value.trim();
    const category = document.getElementById("manualWalletCategory").value.trim();
    const color = document.getElementById("manualWalletColor").value;
    if (!address) { showToast("Enter a wallet address first", true); return; }
    fetch("/known_wallets/category", {
        method: "POST", headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ address, category, color })
    }).then(r => r.json()).then(data => {
        if (data.error) { showToast(data.error, true); return; }
        document.getElementById("manualWalletAddress").value = "";
        document.getElementById("manualWalletCategory").value = "";
        showToast("Wallet added to the database", false);
        if (currentMainTab === "analysis" && currentAnalysisTab === "addresses") loadAddresses(document.getElementById("analysisContent"));
    }).catch(err => showToast("Network error: " + err, true));
}

function renderSuspectFilterList() {
    const list = document.getElementById("suspectFilterList");
    const ids = Object.keys(suspects);
    if (!ids.length) {
        list.innerHTML = '<div style="color:var(--text-dim);font-size:12.5px;">No suspects yet.</div>';
        return;
    }
    list.innerHTML = ids.map(id => `
        <label class="suspect-filter-item">
            <input type="checkbox" ${suspectFilterSelected.has(id) ? "checked" : ""} onchange="onSuspectFilterChange('${id}', this.checked)">
            <span>${escapeHtml(suspects[id])}</span>
        </label>
    `).join("");
}

function onSuspectFilterChange(id, checked) {
    if (checked) suspectFilterSelected.add(id);
    else suspectFilterSelected.delete(id);
    refreshCurrentAnalysisTab();
}

function setAllSuspectFilter(selectAll) {
    const ids = Object.keys(suspects);
    suspectFilterSelected = selectAll ? new Set(ids) : new Set();
    renderSuspectFilterList();
    refreshCurrentAnalysisTab();
}

function refreshCurrentAnalysisTab() {
    if (currentMainTab === "analysis") loadAnalysisTab(currentAnalysisTab);
}

// Builds the ?suspects=id1,id2 query string shared by every analysis fetch and by export,
// so the same suspect-filter selection drives both the on-screen tabs and exports. Empty
// string when nothing is filtered out (i.e. show/export everything).
function suspectsQueryParam() {
    const allIds = Object.keys(suspects);
    const selected = Array.from(suspectFilterSelected);
    const isFiltered = selected.length > 0 && selected.length < allIds.length;
    return isFiltered ? `?suspects=${selected.join(",")}` : "";
}

function triggerExport() {
    const allIds = Object.keys(suspects);
    const selected = Array.from(suspectFilterSelected);

    if (allIds.length && selected.length === 0) {
        showToast("No suspects selected for export - check the filter.", true);
        return;
    }

    window.location.href = `/export/xlsx${suspectsQueryParam()}`;
}


// ---------------------------------------------------------------------------
// CASE SAVE / LOAD
// ---------------------------------------------------------------------------

function saveCase() {
    window.location.href = "/case/export";
}

function loadCase(file) {
    if (!file) return;
    if (!confirm("Loading a case replaces everything currently open (suspects, files, hidden wallets) with the saved case. Continue?")) {
        document.getElementById("caseFileInput").value = "";
        return;
    }

    const formData = new FormData();
    formData.append("file", file);

    fetch("/case/import", { method: "POST", body: formData })
        .then(r => r.json().then(data => ({ ok: r.ok, data })))
        .then(({ ok, data }) => {
            document.getElementById("caseFileInput").value = "";
            if (!ok || data.error) { showToast(data.error || "Failed to load case", true); return; }

            activeSuspectId = "";
            collapsedSuspects = {};
            catalogData = {};
            suspectFilterSelected = new Set();
            renderCaseName(data.case_name);

            fetch("/files").then(r => r.json()).then(files => {
                files.forEach(item => { catalogData[item.id] = item; });
                refreshSuspects();
                updateDropzoneState();
                showToast(`Case loaded: ${data.suspect_count} suspect(s), ${data.file_count} file(s)`, false);
            });
        })
        .catch(err => {
            document.getElementById("caseFileInput").value = "";
            showToast("Network error loading case: " + err, true);
        });
}


// ---------------------------------------------------------------------------
// CASE NAME
// ---------------------------------------------------------------------------

function renderCaseName(name) {
    currentCaseName = name || "";
    const badge = document.getElementById("caseNameDisplay");
    if (currentCaseName) {
        badge.textContent = "📁 " + currentCaseName + " ✏️";
        badge.classList.remove("untitled");
    } else {
        badge.textContent = "📁 Untitled case ✏️";
        badge.classList.add("untitled");
    }
}

function startEditCaseName() {
    const input = document.getElementById("caseNameInput");
    document.getElementById("caseNameDisplay").style.display = "none";
    input.style.display = "inline-block";
    input.value = currentCaseName;
    input.focus();
    input.select();
}

function onCaseNameKeydown(event) {
    if (event.key === "Enter") { event.target.blur(); }
    else if (event.key === "Escape") { event.target.value = currentCaseName; event.target.blur(); }
}

function saveCaseName() {
    const input = document.getElementById("caseNameInput");
    const name = input.value.trim();
    input.style.display = "none";
    document.getElementById("caseNameDisplay").style.display = "inline-flex";
    if (name === currentCaseName) return;
    fetch("/case/name", { method: "POST", headers: { "Content-Type": "application/json" }, body: JSON.stringify({ name }) })
        .then(r => r.json()).then(data => renderCaseName(data.name))
        .catch(err => showToast("Network error saving case name: " + err, true));
}

function cleanAll() {
    if (!confirm("Clean all wipes the current case (suspects, files, hidden wallets, notes, case name). The known-wallets database is kept. Continue?")) return;
    fetch("/case/reset", { method: "POST" })
        .then(r => r.json())
        .then(() => {
            activeSuspectId = "";
            collapsedSuspects = {};
            catalogData = {};
            suspectFilterSelected = new Set();
            renderCaseName("");
            dismissAutosaveBanner();
            refreshSuspects();
            updateDropzoneState();
            showToast("Case cleared.", false);
        })
        .catch(err => showToast("Network error clearing case: " + err, true));
}

// All state lives server-side (SUSPECTS/CATALOG in memory) - a page refresh doesn't lose
// anything on the server, but catalogData is a client-side JS object that starts empty on
// every page load, so without this the Files tab would show "No files imported yet." right
// after a refresh even though the server still has everything.
function initApp() {
    fetch("/case/name").then(r => r.json()).then(data => renderCaseName(data.name));
    fetch("/files").then(r => r.json()).then(files => {
        files.forEach(item => { catalogData[item.id] = item; });
        refreshSuspects();
        // Only offer to restore an autosave into an otherwise-empty session - if files were
        // already reloaded here (server never restarted) there's nothing to recover.
        if (files.length === 0) checkAutosaveBanner();
    });
}

function checkAutosaveBanner() {
    fetch("/case/autosave/status").then(r => r.json()).then(status => {
        if (!status.exists) return;
        const when = status.saved_at ? new Date(status.saved_at).toLocaleString(undefined, { hour12: false }) : "an earlier session";
        const banner = document.getElementById("autosaveBanner");
        banner.innerHTML = `
            <span>💾 Found an autosave from ${escapeHtml(when)} - looks like the app didn't close cleanly last time.</span>
            <span class="actions">
                <button class="btn small" onclick="restoreAutosave()">Restore</button>
                <button class="btn small" onclick="dismissAutosaveBanner()">Dismiss</button>
            </span>
        `;
        banner.style.display = "flex";
    }).catch(() => {});
}

function dismissAutosaveBanner() {
    document.getElementById("autosaveBanner").style.display = "none";
}

function restoreAutosave() {
    fetch("/case/autosave/restore", { method: "POST" })
        .then(r => r.json().then(data => ({ ok: r.ok, data })))
        .then(({ ok, data }) => {
            if (!ok || data.error) { showToast(data.error || "Failed to restore autosave", true); return; }
            dismissAutosaveBanner();
            activeSuspectId = "";
            collapsedSuspects = {};
            catalogData = {};
            suspectFilterSelected = new Set();
            renderCaseName(data.case_name);
            fetch("/files").then(r => r.json()).then(files => {
                files.forEach(item => { catalogData[item.id] = item; });
                refreshSuspects();
                updateDropzoneState();
                showToast(`Autosave restored: ${data.suspect_count} suspect(s), ${data.file_count} file(s)`, false);
            });
        })
        .catch(err => showToast("Network error restoring autosave: " + err, true));
}

// ---------------------------------------------------------------------------
// THEME
// ---------------------------------------------------------------------------

function applyThemeIcon() {
    const theme = document.documentElement.getAttribute("data-theme");
    document.getElementById("themeToggleBtn").textContent = theme === "light" ? "🌙" : "☀️";
}

function toggleTheme() {
    const next = document.documentElement.getAttribute("data-theme") === "light" ? "dark" : "light";
    document.documentElement.setAttribute("data-theme", next);
    localStorage.setItem("cryptolink-theme", next);
    applyThemeIcon();
    // Node label color is baked into the canvas by vis-network, not styled via CSS - only
    // needs a redraw if the Graph tab is the one currently on screen.
    if (currentMainTab === "analysis" && currentAnalysisTab === "graph") {
        loadGraph(document.getElementById("analysisContent"));
    }
}

applyThemeIcon();
initApp();
</script>

</body>
</html>
"""

if __name__ == "__main__":
    print("=" * 50)
    print("CryptoLink starting.")
    print("Open your browser at: http://127.0.0.1:5000")
    print("=" * 50)

    # No reloader/debugger once packaged as a standalone .exe by PyInstaller - there's no
    # source file for it to watch for changes to, and the reloader's re-exec trick doesn't
    # play well with a frozen executable.
    is_frozen = getattr(sys, "frozen", False)
    use_debug = not is_frozen

    # debug=True runs Flask under a reloader, which re-executes this whole block in a child
    # process - only start the background thread/browser-open there (WERKZEUG_RUN_MAIN=="true"),
    # not in the parent watcher process too, or everything would run twice.
    if not use_debug or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        start_autosave_thread()
        atexit.register(_write_autosave)  # best-effort final save on a clean Ctrl+C/exit
        threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5000")).start()

    app.run(debug=use_debug, port=5000)
